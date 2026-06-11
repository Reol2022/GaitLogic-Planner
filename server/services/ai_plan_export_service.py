from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from server.common.exceptions import BadRequestError

SUPPORTED_EXPORT_FORMATS = {
    "xlsx",
    "csv",
    "markdown",
    "md",
    "json",
    "ics",
    "garmin_csv",
    "coros_csv",
    "device_csv",
}


@dataclass(frozen=True)
class AIPlanExportFile:
    content: bytes
    filename: str
    media_type: str


HEADERS = [
    "日期",
    "星期",
    "训练块",
    "阶段",
    "训练内容",
    "重点说明",
    "计划km",
    "主类型",
    "目标配速",
    "排序",
]


def export_ai_plan_draft(draft: Any, export_format: str) -> AIPlanExportFile:
    normalized_format = export_format.lower()
    if normalized_format not in SUPPORTED_EXPORT_FORMATS:
        raise BadRequestError("Unsupported export format.")

    filename_stem = _safe_filename(draft.title or f"ai-plan-{draft.id}")
    if normalized_format == "xlsx":
        return AIPlanExportFile(
            content=_to_xlsx(draft),
            filename=f"{filename_stem}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    if normalized_format == "csv":
        return AIPlanExportFile(
            content=_to_csv(draft),
            filename=f"{filename_stem}.csv",
            media_type="text/csv; charset=utf-8",
        )
    if normalized_format in ("markdown", "md"):
        return AIPlanExportFile(
            content=_to_markdown(draft),
            filename=f"{filename_stem}.md",
            media_type="text/markdown; charset=utf-8",
        )
    if normalized_format == "json":
        return AIPlanExportFile(
            content=_to_json(draft),
            filename=f"{filename_stem}.json",
            media_type="application/json; charset=utf-8",
        )
    if normalized_format == "ics":
        return AIPlanExportFile(
            content=_to_ics(draft),
            filename=f"{filename_stem}.ics",
            media_type="text/calendar; charset=utf-8",
        )
    return AIPlanExportFile(
        content=_to_device_csv(draft, normalized_format),
        filename=f"{filename_stem}-{normalized_format}.csv",
        media_type="text/csv; charset=utf-8",
    )


def _to_xlsx(draft: Any) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "草稿说明"
    summary_rows = [
        ("标题", draft.title),
        ("目标", draft.goal),
        ("开始日期", _date_text(draft.start_date)),
        ("结束日期", _date_text(draft.end_date)),
        ("目标比赛", draft.target_race_name),
        ("目标比赛日期", _date_text(draft.target_race_date)),
        ("目标成绩", draft.target_result),
        ("摘要", draft.summary),
        ("风险提示", "；".join(draft.risk_notes or [])),
        ("导出说明", "设备参考 CSV 仅用于手动录入或二次转换，不代表 Garmin / 高驰官方本地课表导入格式。"),
    ]
    for row in summary_rows:
        summary_sheet.append(row)
    summary_sheet.column_dimensions["A"].width = 18
    summary_sheet.column_dimensions["B"].width = 80
    summary_sheet["A1"].font = Font(bold=True)

    workout_sheet = workbook.create_sheet("训练计划")
    workout_sheet.append(HEADERS)
    for cell in workout_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1976D2")
    for workout in _sorted_workouts(draft):
        workout_sheet.append(_workout_row(workout))
    for index, width in enumerate([14, 10, 20, 18, 44, 32, 10, 12, 18, 8], start=1):
        workout_sheet.column_dimensions[get_column_letter(index)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _to_csv(draft: Any) -> bytes:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(HEADERS)
    for workout in _sorted_workouts(draft):
        writer.writerow(_workout_row(workout))
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _to_device_csv(draft: Any, export_format: str) -> bytes:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "platform_reference",
            "date",
            "weekday",
            "workout_name",
            "sport",
            "planned_distance_km",
            "main_type",
            "target_pace",
            "workout_description",
            "notes",
        ]
    )
    platform = {
        "garmin_csv": "Garmin manual reference",
        "coros_csv": "COROS manual reference",
        "device_csv": "Device manual reference",
    }.get(export_format, "Device manual reference")
    for workout in _sorted_workouts(draft):
        writer.writerow(
            [
                platform,
                _date_text(workout.workout_date),
                workout.weekday or "",
                _workout_name(workout),
                "Run",
                _number_text(workout.planned_distance_km),
                workout.main_type_raw or _enum_value(workout.main_type_normalized),
                workout.target_pace_text or "",
                workout.planned_content,
                workout.focus_note or "",
            ]
        )
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _to_markdown(draft: Any) -> bytes:
    lines = [
        f"# {draft.title}",
        "",
        f"- 目标：{draft.goal or '-'}",
        f"- 日期：{_date_text(draft.start_date)} 至 {_date_text(draft.end_date)}",
        f"- 目标比赛：{draft.target_race_name or '-'}",
        f"- 目标成绩：{draft.target_result or '-'}",
        "",
        draft.summary or "",
        "",
    ]
    if draft.risk_notes:
        lines.extend(["## 风险提示", ""])
        lines.extend(f"- {note}" for note in draft.risk_notes)
        lines.append("")

    lines.extend(
        [
            "## 训练计划",
            "",
            "| 日期 | 星期 | 训练块 | 阶段 | 内容 | km | 类型 | 配速 |",
            "| --- | --- | --- | --- | --- | ---: | --- | --- |",
        ]
    )
    for workout in _sorted_workouts(draft):
        lines.append(
            "| "
            + " | ".join(
                [
                    _date_text(workout.workout_date),
                    workout.weekday or "",
                    workout.block_name or "",
                    workout.phase_name or "",
                    _escape_markdown_table(workout.planned_content),
                    _number_text(workout.planned_distance_km),
                    workout.main_type_raw or _enum_value(workout.main_type_normalized),
                    workout.target_pace_text or "",
                ]
            )
            + " |"
        )
    return "\n".join(lines).encode("utf-8")


def _to_json(draft: Any) -> bytes:
    payload = {
        "id": draft.id,
        "title": draft.title,
        "goal": draft.goal,
        "start_date": _date_text(draft.start_date),
        "end_date": _date_text(draft.end_date),
        "target_race_name": draft.target_race_name,
        "target_race_date": _date_text(draft.target_race_date),
        "target_result": draft.target_result,
        "summary": draft.summary,
        "risk_notes": draft.risk_notes or [],
        "workouts": [
            {
                "workout_date": _date_text(workout.workout_date),
                "weekday": workout.weekday,
                "block_name": workout.block_name,
                "phase_name": workout.phase_name,
                "planned_content": workout.planned_content,
                "focus_note": workout.focus_note,
                "planned_distance_km": _number_value(workout.planned_distance_km),
                "main_type_raw": workout.main_type_raw,
                "main_type_normalized": _enum_value(workout.main_type_normalized),
                "target_pace_text": workout.target_pace_text,
                "sort_order": workout.sort_order,
            }
            for workout in _sorted_workouts(draft)
        ],
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def _to_ics(draft: Any) -> bytes:
    now = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//GaitLogic Planner//AI Plan Draft//CN",
        "CALSCALE:GREGORIAN",
    ]
    for workout in _sorted_workouts(draft):
        date_value = workout.workout_date.strftime("%Y%m%d")
        summary = _ics_escape(_workout_name(workout))
        description = _ics_escape(
            "\n".join(
                value
                for value in [
                    workout.planned_content,
                    f"重点：{workout.focus_note}" if workout.focus_note else "",
                    f"计划距离：{_number_text(workout.planned_distance_km)} km" if workout.planned_distance_km else "",
                    f"目标配速：{workout.target_pace_text}" if workout.target_pace_text else "",
                    "来自 GaitLogic Planner AI 课表草稿。",
                ]
                if value
            )
        )
        lines.extend(
            [
                "BEGIN:VEVENT",
                f"UID:ai-plan-{draft.id}-{workout.sort_order}@gaitlogic-planner",
                f"DTSTAMP:{now}",
                f"DTSTART;VALUE=DATE:{date_value}",
                f"SUMMARY:{summary}",
                f"DESCRIPTION:{description}",
                "END:VEVENT",
            ]
        )
    lines.append("END:VCALENDAR")
    return "\r\n".join(lines).encode("utf-8")


def _workout_row(workout: Any) -> list[str]:
    return [
        _date_text(workout.workout_date),
        workout.weekday or "",
        workout.block_name or "",
        workout.phase_name or "",
        workout.planned_content,
        workout.focus_note or "",
        _number_text(workout.planned_distance_km),
        workout.main_type_raw or _enum_value(workout.main_type_normalized),
        workout.target_pace_text or "",
        str(workout.sort_order),
    ]


def _sorted_workouts(draft: Any) -> list[Any]:
    return sorted(draft.workouts or [], key=lambda item: (item.workout_date, item.sort_order))


def _workout_name(workout: Any) -> str:
    type_text = workout.main_type_raw or _enum_value(workout.main_type_normalized)
    distance = f" { _number_text(workout.planned_distance_km) }km" if workout.planned_distance_km else ""
    return f"{type_text}{distance}".strip()


def _date_text(value: Any) -> str:
    return value.isoformat() if value else ""


def _number_text(value: Any) -> str:
    if value is None:
        return ""
    return f"{float(value):g}"


def _number_value(value: Any) -> float | None:
    if value is None:
        return None
    return float(value)


def _enum_value(value: Any) -> str:
    return str(getattr(value, "value", value))


def _safe_filename(value: str) -> str:
    cleaned = "".join(char if char.isalnum() or char in ("-", "_") else "-" for char in value.strip())
    return cleaned.strip("-")[:80] or "ai-plan-draft"


def _escape_markdown_table(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", "<br>")


def _ics_escape(value: str) -> str:
    return value.replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace("\n", "\\n")
