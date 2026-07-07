from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

from planner_core.database.models import (
    PlanAdjustmentDraft,
    PlanAdjustmentItem,
    PlanImportAudit,
    PlannedWorkout,
    TrainingBlock,
    TrainingCycle,
    WorkoutLog,
)
from planner_core.enums import PlanAdjustmentAction, PlanAdjustmentDraftStatus, WorkoutMainTypeNormalized, WorkoutStatusNormalized
from planner_core.utils.excel_parse import normalize_workout_main_type, parse_date, parse_decimal
from server.common.exceptions import BadRequestError, NotFoundError
from server.schemas.plan_import import (
    PlanImportApplyResponse,
    PlanImportCreateResponse,
    PlanImportDiffSummary,
    PlanImportDraftRead,
    PlanImportIssue,
    PlanImportItemUpdate,
    PlanImportStructuredRequest,
    PlanImportWorkoutItem,
)
from server.services.training_cycle_lifecycle_service import get_active_cycle
from server.services.weekly_review_stats_service import COMPLETED_STATUSES, local_today

PARSER_VERSION = "plan-import-parser-v1"
MAX_FILE_SIZE_BYTES = 2 * 1024 * 1024
MAX_ROWS = 370
SUPPORTED_EXTENSIONS = {".json", ".xlsx", ".csv", ".txt", ".md"}
TEXT_LINE_PATTERN = re.compile(r"^(DAY\+(?P<offset>\d+)|(?P<date>\d{4}-\d{2}-\d{2}))$")


def create_structured_import(
    db: Session,
    user_id: int,
    payload: PlanImportStructuredRequest,
    idempotency_key: str | None = None,
) -> PlanImportCreateResponse:
    request_id = idempotency_key or payload.client_request_id
    existing = _find_existing_draft(db, user_id, request_id)
    if existing:
        return _create_response(existing)
    return _create_import_draft(db, user_id, payload, "structured_json", None, _hash_payload(payload.model_dump(mode="json")))


async def create_file_import(
    db: Session,
    user_id: int,
    *,
    filename: str,
    content_type: str | None,
    content: bytes,
    target_cycle_id: int | None,
    target_block_id: int | None,
    client_request_id: str,
    anchor_strategy: str,
    effective_date: date | None,
    merge_strategy: str,
    timezone: str,
    idempotency_key: str | None = None,
) -> PlanImportCreateResponse:
    if not content:
        raise BadRequestError("上传文件不能为空。")
    if len(content) > MAX_FILE_SIZE_BYTES:
        raise BadRequestError("文件过大，单次导入最大支持 2MB。")
    safe_name = filename.rsplit("\\", 1)[-1].rsplit("/", 1)[-1]
    extension = "." + safe_name.rsplit(".", 1)[-1].lower() if "." in safe_name else ""
    if extension == ".xls":
        raise BadRequestError("暂不支持 .xls，请另存为 .xlsx 后再上传。")
    if extension not in SUPPORTED_EXTENSIONS:
        raise BadRequestError("不支持的文件格式。")
    _validate_mime(extension, content_type)
    request_id = idempotency_key or client_request_id
    existing = _find_existing_draft(db, user_id, request_id)
    if existing:
        return _create_response(existing)

    workouts = _parse_file(extension, content)
    payload = PlanImportStructuredRequest(
        target_cycle_id=target_cycle_id,
        target_block_id=target_block_id,
        source=safe_name,
        client_request_id=client_request_id,
        anchor_strategy=anchor_strategy,  # type: ignore[arg-type]
        effective_date=effective_date,
        merge_strategy=merge_strategy,  # type: ignore[arg-type]
        timezone=timezone,
        workouts=workouts,
    )
    return _create_import_draft(db, user_id, payload, _source_type_for_extension(extension), safe_name, hashlib.sha256(content).hexdigest())


def get_plan_import(db: Session, user_id: int, import_id: int) -> PlanImportDraftRead:
    return PlanImportDraftRead.model_validate(_get_draft(db, user_id, import_id))


def update_plan_import_item(
    db: Session, user_id: int, import_id: int, item_id: int, payload: PlanImportItemUpdate
) -> PlanImportDraftRead:
    draft = _get_draft(db, user_id, import_id)
    if draft.status in {PlanAdjustmentDraftStatus.applied, PlanAdjustmentDraftStatus.cancelled, PlanAdjustmentDraftStatus.expired}:
        raise BadRequestError("该导入草稿已不可编辑。")
    item = next((candidate for candidate in draft.items if candidate.id == item_id), None)
    if item is None:
        raise NotFoundError("草稿条目不存在。")
    data = dict(item.normalized_item_json or {})
    update = payload.model_dump(exclude_unset=True, mode="json")
    if "workout_type" in update and update["workout_type"] is not None:
        update["workout_type"] = update["workout_type"]
    data.update({key: value for key, value in update.items() if value is not None})
    normalized = PlanImportWorkoutItem.model_validate(data)
    item.normalized_item_json = normalized.model_dump(mode="json")
    db.commit()
    return validate_plan_import(db, user_id, import_id)


def validate_plan_import(db: Session, user_id: int, import_id: int) -> PlanImportDraftRead:
    draft = _get_draft(db, user_id, import_id)
    payload = _payload_from_draft(draft)
    diff = _build_diff(db, user_id, payload, draft.source_type or "structured_json")
    _replace_draft_items(draft, diff)
    db.commit()
    return PlanImportDraftRead.model_validate(_get_draft(db, user_id, import_id))


def apply_plan_import(db: Session, user_id: int, import_id: int) -> PlanImportApplyResponse:
    try:
        draft = _get_draft(db, user_id, import_id, lock=True)
        if draft.status == PlanAdjustmentDraftStatus.applied:
            return PlanImportApplyResponse(
                import_id=draft.id,
                status=draft.status,
                diff_summary=PlanImportDiffSummary.model_validate(draft.diff_summary_json or {}),
            )
        if draft.status in {PlanAdjustmentDraftStatus.cancelled, PlanAdjustmentDraftStatus.expired}:
            raise BadRequestError("该导入草稿已不可应用。")
        payload = _payload_from_draft(draft)
        diff = _build_diff(db, user_id, payload, draft.source_type or "structured_json")
        if diff["conflicts"]:
            _replace_draft_items(draft, diff)
            db.commit()
            raise BadRequestError("导入草稿仍存在冲突，不能应用。")

        for item in diff["items"]:
            operation = item["operation"]
            existing = item.get("existing")
            normalized = item.get("normalized")
            if operation == "remove" and existing:
                db.delete(existing)
            elif operation == "update" and existing and normalized:
                _apply_normalized_to_workout(existing, normalized)
                existing.plan_version += 1
            elif operation == "create" and normalized:
                workout = PlannedWorkout(
                    user_id=user_id,
                    cycle_id=payload.target_cycle_id,
                    block_id=item["block"].id,
                    workout_date=normalized.planned_date,
                    session_index=normalized.session_index,
                    date_text=normalized.planned_date.isoformat() if normalized.planned_date else None,
                    planned_content=normalized.content,
                    focus_note=normalized.notes,
                    target_pace_text=normalized.target_pace,
                    planned_distance_km=normalized.planned_distance_km or Decimal("0"),
                    main_type_raw=normalized.workout_type,
                    main_type_normalized=_normalize_type(normalized),
                    source_sheet="plan_import",
                    source_row=None,
                    sort_order=_sort_order(normalized),
                )
                db.add(workout)
        _replace_draft_items(draft, diff)
        draft.status = PlanAdjustmentDraftStatus.applied
        draft.applied_at = datetime.utcnow()
        summary = diff["summary"]
        db.add(
            PlanImportAudit(
                user_id=user_id,
                import_id=draft.id,
                source_type=draft.source_type,
                merge_strategy=draft.merge_strategy,
                effective_date=draft.effective_date,
                created_count=summary.created_count,
                updated_count=summary.updated_count,
                removed_count=summary.removed_count,
                protected_count=summary.protected_count,
                applied_at=draft.applied_at,
                actor_type="user",
                client_request_id=draft.client_request_id,
            )
        )
        db.commit()
        return PlanImportApplyResponse(
            import_id=draft.id,
            status=draft.status,
            diff_summary=PlanImportDiffSummary.model_validate(draft.diff_summary_json or {}),
        )
    except Exception:
        db.rollback()
        raise


def cancel_plan_import(db: Session, user_id: int, import_id: int) -> PlanImportDraftRead:
    draft = _get_draft(db, user_id, import_id)
    if draft.status == PlanAdjustmentDraftStatus.applied:
        raise BadRequestError("已应用的导入草稿不能取消。")
    draft.status = PlanAdjustmentDraftStatus.cancelled
    draft.cancelled_at = datetime.utcnow()
    db.commit()
    return PlanImportDraftRead.model_validate(draft)


def _find_existing_draft(db: Session, user_id: int, request_id: str | None) -> PlanAdjustmentDraft | None:
    if not request_id:
        return None
    return db.scalar(
        select(PlanAdjustmentDraft)
        .options(selectinload(PlanAdjustmentDraft.items))
        .where(PlanAdjustmentDraft.user_id == user_id, PlanAdjustmentDraft.client_request_id == request_id)
    )


def _get_draft(db: Session, user_id: int, import_id: int, *, lock: bool = False) -> PlanAdjustmentDraft:
    stmt = (
        select(PlanAdjustmentDraft)
        .options(selectinload(PlanAdjustmentDraft.items).selectinload(PlanAdjustmentItem.planned_workout))
        .where(PlanAdjustmentDraft.id == import_id, PlanAdjustmentDraft.user_id == user_id)
    )
    if lock:
        stmt = stmt.with_for_update()
    draft = db.scalar(stmt)
    if draft is None:
        raise NotFoundError("导入草稿不存在。")
    draft.items.sort(key=lambda item: (item.planned_date or date.max, item.session_index or 99, item.id))
    return draft


def _create_import_draft(
    db: Session,
    user_id: int,
    payload: PlanImportStructuredRequest,
    source_type: str,
    source_filename: str | None,
    raw_payload_hash: str,
) -> PlanImportCreateResponse:
    diff = _build_diff(db, user_id, payload, source_type)
    draft = PlanAdjustmentDraft(
        user_id=user_id,
        review_report_id=None,
        cycle_id=payload.target_cycle_id or diff["cycle"].id,
        source_block_id=None,
        target_block_id=payload.target_block_id,
        target_cycle_id=diff["cycle"].id,
        status=PlanAdjustmentDraftStatus.conflict if diff["conflicts"] else PlanAdjustmentDraftStatus.ready,
        summary="外部课表导入草稿",
        source_type=source_type,
        source_name=payload.source,
        source_filename=source_filename,
        raw_payload_hash=raw_payload_hash,
        parser_version=PARSER_VERSION,
        merge_strategy=payload.merge_strategy,
        anchor_strategy=payload.anchor_strategy,
        effective_date=diff["effective_date"],
        normalized_payload_json=[item.model_dump(mode="json") for item in diff["normalized"]],
        diff_summary_json=diff["summary"].model_dump(mode="json"),
        conflict_summary_json=[issue.model_dump(mode="json") for issue in diff["conflicts"]],
        warnings_json=[issue.model_dump(mode="json") for issue in diff["warnings"]],
        client_request_id=payload.client_request_id,
        expires_at=datetime.utcnow() + timedelta(days=14),
    )
    _replace_draft_items(draft, diff)
    db.add(draft)
    db.commit()
    db.refresh(draft)
    return _create_response(draft)


def _create_response(draft: PlanAdjustmentDraft) -> PlanImportCreateResponse:
    return PlanImportCreateResponse(
        import_id=draft.id,
        status=draft.status,
        effective_date=draft.effective_date or local_today(),
        normalized_items=draft.normalized_payload_json or [],
        diff_summary=PlanImportDiffSummary.model_validate(draft.diff_summary_json or {}),
        conflicts=[PlanImportIssue.model_validate(item) for item in draft.conflict_summary_json or []],
        warnings=[PlanImportIssue.model_validate(item) for item in draft.warnings_json or []],
    )


def _payload_from_draft(draft: PlanAdjustmentDraft) -> PlanImportStructuredRequest:
    return PlanImportStructuredRequest(
        target_cycle_id=draft.target_cycle_id or draft.cycle_id,
        target_block_id=draft.target_block_id,
        source=draft.source_name or "draft",
        client_request_id=draft.client_request_id or f"draft-{draft.id}",
        anchor_strategy=(draft.anchor_strategy or "after_last_completed"),  # type: ignore[arg-type]
        effective_date=draft.effective_date,
        merge_strategy=(draft.merge_strategy or "replace_uncompleted_in_range"),  # type: ignore[arg-type]
        timezone="Asia/Shanghai",
        workouts=[PlanImportWorkoutItem.model_validate(item) for item in draft.normalized_payload_json or []],
    )


def _build_diff(db: Session, user_id: int, payload: PlanImportStructuredRequest, source_type: str) -> dict[str, Any]:
    cycle = _resolve_cycle(db, user_id, payload.target_cycle_id, payload.effective_date)
    effective_date = _resolve_effective_date(db, user_id, cycle.id, payload)
    normalized, validation_issues = _normalize_items(payload.workouts, effective_date)
    existing = _existing_workouts(db, user_id, cycle.id)
    blocks = _blocks_for_cycle(db, user_id, cycle.id)
    by_key = {(item.workout_date, item.session_index): item for item in existing if item.workout_date}
    incoming_keys: set[tuple[date, int]] = set()
    items: list[dict[str, Any]] = []
    conflicts: list[PlanImportIssue] = list(validation_issues)
    warnings: list[PlanImportIssue] = []
    summary = PlanImportDiffSummary()

    for normalized_item in normalized:
        key = (normalized_item.planned_date, normalized_item.session_index)
        if key in incoming_keys:
            conflicts.append(_issue("DUPLICATE_SESSION", "同一天 session_index 重复。", normalized_item))
            continue
        incoming_keys.add(key)
        block = _block_for_date(blocks, normalized_item.planned_date, payload.target_block_id)
        if block is None:
            conflicts.append(_issue("NO_MATCHING_BLOCK", "目标日期不属于任何训练块。", normalized_item))
            continue
        existing_workout = by_key.get(key)
        if existing_workout:
            if _is_protected(existing_workout):
                summary.protected_count += 1
                conflicts.append(_issue("PROTECTED_WORKOUT", "目标日期已有受保护计划，不能覆盖。", normalized_item))
                items.append(_item("conflict", normalized_item, existing_workout, block))
            elif payload.merge_strategy == "fill_empty_only":
                summary.preserved_count += 1
                warnings.append(_issue("EXISTING_WORKOUT_PRESERVED", "fill_empty_only 策略不会修改已有计划。", normalized_item))
                items.append(_item("preserve", normalized_item, existing_workout, block))
            else:
                summary.updated_count += 1
                items.append(_item("update", normalized_item, existing_workout, block))
        else:
            summary.created_count += 1
            items.append(_item("create", normalized_item, None, block))

    for existing_workout in existing:
        if not existing_workout.workout_date or (existing_workout.workout_date, existing_workout.session_index) in incoming_keys:
            continue
        if _should_remove(existing_workout, normalized, effective_date, payload.merge_strategy):
            if _is_protected(existing_workout):
                summary.protected_count += 1
                summary.preserved_count += 1
                continue
            summary.removed_count += 1
            items.append(_item("remove", None, existing_workout, existing_workout.block))
        else:
            summary.preserved_count += 1

    summary.conflict_count = len(conflicts)
    summary.warning_count = len(warnings)
    return {
        "cycle": cycle,
        "effective_date": effective_date,
        "normalized": normalized,
        "summary": summary,
        "conflicts": conflicts,
        "warnings": warnings,
        "items": items,
    }


def _normalize_items(items: list[PlanImportWorkoutItem], effective_date: date) -> tuple[list[PlanImportWorkoutItem], list[PlanImportIssue]]:
    normalized: list[PlanImportWorkoutItem] = []
    issues: list[PlanImportIssue] = []
    for item in items[:MAX_ROWS]:
        data = item.model_dump()
        computed = effective_date + timedelta(days=item.day_offset or 0) if item.day_offset is not None else item.planned_date
        if item.planned_date and item.day_offset is not None and computed != item.planned_date:
            issues.append(_issue("DATE_OFFSET_MISMATCH", "planned_date 与 day_offset 计算结果不一致。", item))
            continue
        data["planned_date"] = computed
        data["workout_type"] = "REST" if item.is_rest_day else item.workout_type
        if item.is_rest_day and data.get("planned_distance_km") is None:
            data["planned_distance_km"] = Decimal("0")
        normalized.append(PlanImportWorkoutItem.model_validate(data))
    return normalized, issues


def _resolve_cycle(db: Session, user_id: int, target_cycle_id: int | None, effective_date: date | None) -> TrainingCycle:
    if target_cycle_id:
        cycle = db.scalar(select(TrainingCycle).where(TrainingCycle.id == target_cycle_id, TrainingCycle.user_id == user_id))
        if cycle is None:
            raise NotFoundError("目标训练周期不存在。")
        return cycle
    target = effective_date or local_today()
    cycle = db.scalar(
        select(TrainingCycle)
        .where(TrainingCycle.user_id == user_id, TrainingCycle.start_date <= target, TrainingCycle.end_date >= target)
        .order_by(TrainingCycle.start_date.desc())
    )
    if cycle is None:
        cycle = db.scalar(select(TrainingCycle).where(TrainingCycle.user_id == user_id).order_by(TrainingCycle.start_date.desc()))
    if cycle is None:
        raise BadRequestError("请先创建训练周期。")
    return cycle


def _resolve_cycle(db: Session, user_id: int, target_cycle_id: int | None, effective_date: date | None) -> TrainingCycle:
    if target_cycle_id:
        cycle = db.scalar(select(TrainingCycle).where(TrainingCycle.id == target_cycle_id, TrainingCycle.user_id == user_id))
        if cycle is None:
            raise NotFoundError("Target training cycle not found.")
        return cycle
    cycle = get_active_cycle(db, user_id)
    if cycle is None:
        raise BadRequestError("当前没有生效中的训练周期，请先启用一个周期或显式选择目标草稿周期。")
    return cycle


def _resolve_effective_date(db: Session, user_id: int, cycle_id: int, payload: PlanImportStructuredRequest) -> date:
    if payload.anchor_strategy == "explicit_date":
        if payload.effective_date is None:
            raise BadRequestError("explicit_date 策略必须提供 effective_date。")
        return payload.effective_date
    last_completed = db.scalar(
        select(func.max(PlannedWorkout.workout_date))
        .join(WorkoutLog, WorkoutLog.planned_workout_id == PlannedWorkout.id)
        .where(
            PlannedWorkout.user_id == user_id,
            PlannedWorkout.cycle_id == cycle_id,
            PlannedWorkout.workout_date.is_not(None),
            WorkoutLog.status_normalized.in_(COMPLETED_STATUSES),
        )
    )
    if last_completed is None:
        return max(local_today(), payload.effective_date or local_today())
    return max(local_today(), last_completed + timedelta(days=1))


def _existing_workouts(db: Session, user_id: int, cycle_id: int) -> list[PlannedWorkout]:
    return list(
        db.scalars(
            select(PlannedWorkout)
            .options(selectinload(PlannedWorkout.workout_log), selectinload(PlannedWorkout.block))
            .where(PlannedWorkout.user_id == user_id, PlannedWorkout.cycle_id == cycle_id)
            .order_by(PlannedWorkout.workout_date, PlannedWorkout.session_index, PlannedWorkout.sort_order)
        )
    )


def _blocks_for_cycle(db: Session, user_id: int, cycle_id: int) -> list[TrainingBlock]:
    return list(
        db.scalars(
            select(TrainingBlock)
            .where(TrainingBlock.user_id == user_id, TrainingBlock.cycle_id == cycle_id)
            .order_by(TrainingBlock.start_date, TrainingBlock.sort_order)
        )
    )


def _block_for_date(blocks: list[TrainingBlock], target_date: date | None, target_block_id: int | None) -> TrainingBlock | None:
    if target_date is None:
        return None
    for block in blocks:
        if target_block_id is not None and block.id != target_block_id:
            continue
        if block.start_date and block.end_date and block.start_date <= target_date <= block.end_date:
            return block
    return None


def _is_protected(workout: PlannedWorkout) -> bool:
    log = workout.workout_log
    return bool(
        workout.is_locked
        or log
        and (
            log.status_normalized in COMPLETED_STATUSES
            or log.actual_distance_km is not None
            or log.actual_duration_seconds is not None
        )
    )


def _should_remove(workout: PlannedWorkout, normalized: list[PlanImportWorkoutItem], effective_date: date, strategy: str) -> bool:
    if workout.workout_date is None:
        return False
    if strategy == "replace_uncompleted_from_date":
        return workout.workout_date >= effective_date
    if strategy == "replace_uncompleted_in_range" and normalized:
        dates = [item.planned_date for item in normalized if item.planned_date]
        return min(dates) <= workout.workout_date <= max(dates)
    return False


def _replace_draft_items(draft: PlanAdjustmentDraft, diff: dict[str, Any]) -> None:
    draft.items.clear()
    draft.normalized_payload_json = [item.model_dump(mode="json") for item in diff["normalized"]]
    draft.diff_summary_json = diff["summary"].model_dump(mode="json")
    draft.conflict_summary_json = [issue.model_dump(mode="json") for issue in diff["conflicts"]]
    draft.warnings_json = [issue.model_dump(mode="json") for issue in diff["warnings"]]
    draft.status = PlanAdjustmentDraftStatus.conflict if diff["conflicts"] else PlanAdjustmentDraftStatus.ready
    for entry in diff["items"]:
        normalized = entry.get("normalized")
        existing = entry.get("existing")
        operation = entry["operation"]
        draft.items.append(
            PlanAdjustmentItem(
                planned_workout_id=existing.id if existing else None,
                operation=operation,
                planned_date=(normalized.planned_date if normalized else existing.workout_date),
                session_index=(normalized.session_index if normalized else existing.session_index),
                action=PlanAdjustmentAction.replace if operation in {"create", "update"} else PlanAdjustmentAction.keep,
                original_content=existing.planned_content if existing else "",
                suggested_content=normalized.content if normalized else existing.planned_content,
                original_distance_km=existing.planned_distance_km if existing else None,
                suggested_distance_km=(normalized.planned_distance_km if normalized else existing.planned_distance_km),
                original_main_type=existing.main_type_normalized.value if existing else None,
                suggested_main_type=_normalize_type(normalized).value if normalized else existing.main_type_normalized.value,
                original_target_pace_text=existing.target_pace_text if existing else None,
                suggested_target_pace_text=normalized.target_pace if normalized else existing.target_pace_text,
                reason=_reason_for_operation(operation),
                is_selected=operation in {"create", "update", "remove"},
                normalized_item_json=normalized.model_dump(mode="json") if normalized else None,
                conflict_json=[],
                warnings_json=[],
                base_plan_version=existing.plan_version if existing else None,
                base_workout_updated_at=existing.updated_at if existing else None,
            )
        )


def _item(operation: str, normalized: PlanImportWorkoutItem | None, existing: PlannedWorkout | None, block: TrainingBlock) -> dict[str, Any]:
    return {"operation": operation, "normalized": normalized, "existing": existing, "block": block}


def _issue(code: str, message: str, item: PlanImportWorkoutItem) -> PlanImportIssue:
    return PlanImportIssue(code=code, message=message, planned_date=item.planned_date, session_index=item.session_index)


def _reason_for_operation(operation: str) -> str:
    return {
        "create": "导入课表新增训练。",
        "update": "导入课表更新未来未完成训练。",
        "remove": "合并策略移除未来未完成训练。",
        "preserve": "计划被保留。",
        "conflict": "存在冲突，不能应用。",
    }.get(operation, "导入课表调整。")


def _normalize_type(item: PlanImportWorkoutItem) -> WorkoutMainTypeNormalized:
    if item.is_rest_day or item.workout_type.upper() == "REST":
        return WorkoutMainTypeNormalized.rest
    return normalize_workout_main_type(item.workout_type)


def _apply_normalized_to_workout(workout: PlannedWorkout, item: PlanImportWorkoutItem) -> None:
    workout.block_id = workout.block_id
    workout.planned_content = item.content
    workout.focus_note = item.notes
    workout.target_pace_text = item.target_pace
    workout.planned_distance_km = item.planned_distance_km or Decimal("0")
    workout.main_type_raw = item.workout_type
    workout.main_type_normalized = _normalize_type(item)
    workout.source_sheet = "plan_import"
    workout.sort_order = _sort_order(item)


def _sort_order(item: PlanImportWorkoutItem) -> int:
    return int(item.planned_date.strftime("%Y%m%d")) * 10 + item.session_index if item.planned_date else item.session_index


def _hash_payload(payload: Any) -> str:
    return hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")).hexdigest()


def _validate_mime(extension: str, content_type: str | None) -> None:
    allowed = {
        ".json": {"application/json", "text/plain", "application/octet-stream"},
        ".xlsx": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/octet-stream"},
        ".csv": {"text/csv", "text/plain", "application/vnd.ms-excel", "application/octet-stream"},
        ".txt": {"text/plain", "application/octet-stream"},
        ".md": {"text/markdown", "text/plain", "application/octet-stream"},
    }
    if content_type and content_type.split(";")[0] not in allowed[extension]:
        raise BadRequestError("文件 MIME 类型与扩展名不匹配。")


def _source_type_for_extension(extension: str) -> str:
    return {".json": "json_file", ".xlsx": "xlsx", ".csv": "csv", ".txt": "txt", ".md": "markdown"}[extension]


def _parse_file(extension: str, content: bytes) -> list[PlanImportWorkoutItem]:
    if extension == ".json":
        try:
            body = json.loads(content.decode("utf-8-sig"))
        except json.JSONDecodeError as exc:
            raise BadRequestError(f"JSON 格式错误：第 {exc.lineno} 行。") from exc
        rows = body.get("workouts", body) if isinstance(body, dict) else body
        return [PlanImportWorkoutItem.model_validate(item) for item in rows]
    if extension == ".xlsx":
        return _parse_xlsx(content)
    if extension == ".csv":
        return _parse_csv(content.decode("utf-8-sig").splitlines())
    return _parse_text(content.decode("utf-8-sig").splitlines())


def _parse_xlsx(content: bytes) -> list[PlanImportWorkoutItem]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    return _rows_to_items(
        (dict(zip(headers, [cell.value for cell in row], strict=False)) for row in sheet.iter_rows(min_row=2)),
        source="xlsx",
    )


def _parse_csv(lines: list[str]) -> list[PlanImportWorkoutItem]:
    return _rows_to_items(csv.DictReader(lines), source="csv")


def _rows_to_items(rows: Any, source: str) -> list[PlanImportWorkoutItem]:
    items: list[PlanImportWorkoutItem] = []
    for row_number, row in enumerate(rows, start=2):
        if row_number > MAX_ROWS + 1:
            raise BadRequestError("导入行数超过限制。")
        if not any(str(value or "").strip() for value in row.values()):
            continue
        try:
            items.append(
                PlanImportWorkoutItem(
                    planned_date=parse_date(row.get("date") or row.get("planned_date")),
                    day_offset=_int_or_none(row.get("day_offset")),
                    session_index=_int_or_none(row.get("session_index")) or 1,
                    workout_type=str(row.get("workout_type") or "").strip(),
                    title=_text_or_none(row.get("title")),
                    planned_distance_km=parse_decimal(row.get("distance_km")),
                    planned_duration_minutes=_int_or_none(row.get("duration_min")),
                    target_pace=_text_or_none(row.get("target_pace")),
                    target_rpe=_int_or_none(row.get("target_rpe")),
                    content=str(row.get("content") or row.get("title") or "").strip(),
                    notes=_text_or_none(row.get("notes")),
                    is_rest_day=str(row.get("is_rest_day") or "").strip().lower() in {"1", "true", "yes", "rest"},
                )
            )
        except (ValidationError, ValueError) as exc:
            raise BadRequestError(f"{source} 第 {row_number} 行格式错误：{exc}") from exc
    if not items:
        raise BadRequestError("文件中没有可导入的训练。")
    return items


def _parse_text(lines: list[str]) -> list[PlanImportWorkoutItem]:
    items: list[PlanImportWorkoutItem] = []
    for row_number, raw in enumerate(lines, start=1):
        line = raw.strip().strip("|")
        if not line or line.startswith("#"):
            continue
        parts = [part.strip() for part in line.split("|")]
        if len(parts) < 5:
            raise BadRequestError(f"文本第 {row_number} 行格式错误。")
        match = TEXT_LINE_PATTERN.match(parts[0])
        if not match:
            raise BadRequestError(f"文本第 {row_number} 行日期格式错误。")
        distance = Decimal("0")
        distance_match = re.search(r"(\d+(?:\.\d+)?)\s*km", parts[2], re.I)
        if distance_match:
            distance = Decimal(distance_match.group(1))
        items.append(
            PlanImportWorkoutItem(
                planned_date=parse_date(match.group("date")) if match.group("date") else None,
                day_offset=int(match.group("offset")) if match.group("offset") else None,
                workout_type=parts[1],
                title=parts[4],
                planned_distance_km=distance,
                target_pace=None if parts[3] == "-" else parts[3],
                content=parts[4],
                notes=parts[5] if len(parts) > 5 and parts[5] else None,
                is_rest_day=parts[1].upper() == "REST",
            )
        )
    if not items:
        raise BadRequestError("文件中没有可导入的训练。")
    return items


def _text_or_none(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None


def _int_or_none(value: Any) -> int | None:
    if value in (None, ""):
        return None
    return int(value)
