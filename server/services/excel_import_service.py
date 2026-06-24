from __future__ import annotations

import hashlib
import json
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from planner_core.database.models import (
    BlockReview,
    ExcelImportJob,
    PaceRule,
    PlannedWorkout,
    TrainingBlock,
    TrainingCycle,
    WorkoutLog,
)
from planner_core.enums import BlockType, ExcelImportStatus, PainScaleVersion, WorkoutStatusNormalized
from planner_core.utils.excel_parse import (
    is_blank,
    normalize_workout_main_type,
    normalize_workout_status,
    parse_date,
    parse_decimal,
    parse_duration_seconds,
    parse_pace_seconds_per_km,
)
from server.schemas.excel_import import ExcelImportErrorItem, ExcelImportResult
from server.services.excel_template_service import (
    SHEET_BLOCKS,
    SHEET_CYCLE,
    SHEET_HEADERS,
    SHEET_LOGS,
    SHEET_PACE_RULES,
    SHEET_REVIEWS,
    SHEET_WORKOUTS,
    STANDARD_SHEETS,
)


def import_excel_workbook(
    db: Session,
    content: bytes,
    file_name: str,
    user_id: int,
) -> ExcelImportResult:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception:
        return _failed_template_result("文件不是有效的 xlsx。")

    template_errors = _validate_template(workbook)
    if template_errors:
        return ExcelImportResult(
            status="failed",
            message="Excel 模板格式错误",
            total_count=0,
            success_count=0,
            failed_count=0,
            errors=template_errors,
        )

    errors: list[ExcelImportErrorItem] = []
    stats = {"total": 0, "success": 0}
    started_at = datetime.now()

    try:
        cycle = _import_cycle(db, workbook[SHEET_CYCLE], user_id, errors, stats)
        cycle_year = cycle.start_date.year if cycle and cycle.start_date else None
        blocks = _import_blocks(db, workbook[SHEET_BLOCKS], user_id, cycle, cycle_year, errors, stats)
        workouts = _import_workouts(
            db,
            workbook[SHEET_WORKOUTS],
            user_id,
            cycle,
            blocks,
            cycle_year,
            errors,
            stats,
        )
        _import_logs(db, workbook[SHEET_LOGS], user_id, workouts, cycle_year, errors, stats)
        _import_reviews(db, workbook[SHEET_REVIEWS], user_id, blocks, errors, stats)
        _import_pace_rules(db, workbook[SHEET_PACE_RULES], user_id, errors, stats)

        failed_count = len(errors)
        status = _result_status(stats["success"], failed_count)
        result = ExcelImportResult(
            status=status,
            message="导入完成" if status != "failed" else "导入失败",
            total_count=stats["total"],
            success_count=stats["success"],
            failed_count=failed_count,
            errors=errors,
        )
        db.add(
            ExcelImportJob(
                user_id=user_id,
                file_name=file_name,
                file_hash=hashlib.sha256(content).hexdigest(),
                sheet_names=workbook.sheetnames,
                status=_job_status(status),
                total_count=result.total_count,
                success_count=result.success_count,
                failed_count=result.failed_count,
                error_message=json.dumps([error.model_dump() for error in errors], ensure_ascii=False)
                if errors
                else None,
                started_at=started_at,
                finished_at=datetime.now(),
            )
        )
        db.commit()
        return result
    except Exception:
        db.rollback()
        raise


def _validate_template(workbook: Any) -> list[ExcelImportErrorItem]:
    errors: list[ExcelImportErrorItem] = []
    for sheet_name in STANDARD_SHEETS:
        if sheet_name not in workbook.sheetnames:
            errors.append(ExcelImportErrorItem(sheet=sheet_name, row=0, message="缺少标准 Sheet。"))
            continue
        if sheet_name not in SHEET_HEADERS:
            continue
        sheet = workbook[sheet_name]
        expected_headers = SHEET_HEADERS[sheet_name]
        actual_headers = [
            _cell_text(sheet.cell(row=1, column=index).value)
            for index in range(1, len(expected_headers) + 1)
        ]
        if actual_headers != expected_headers:
            missing = [header for header in expected_headers if header not in actual_headers]
            message = "表头不匹配"
            if missing:
                message += f"，缺少字段：{'、'.join(missing)}"
            errors.append(ExcelImportErrorItem(sheet=sheet_name, row=1, message=message))
    return errors


def _import_cycle(
    db: Session,
    sheet: Any,
    user_id: int,
    errors: list[ExcelImportErrorItem],
    stats: dict[str, int],
) -> TrainingCycle | None:
    for row_index, row in _iter_data_rows(sheet, SHEET_HEADERS[SHEET_CYCLE]):
        stats["total"] += 1
        try:
            name = _required_text(row["周期名称"], "周期名称")
            start_date = parse_date(row["开始日期"])
            cycle = db.scalar(
                select(TrainingCycle).where(
                    TrainingCycle.user_id == user_id,
                    TrainingCycle.name == name,
                    TrainingCycle.start_date == start_date,
                )
            )
            if cycle is None:
                cycle = TrainingCycle(user_id=user_id, name=name, start_date=start_date)
                db.add(cycle)
            cycle.goal = _optional_text(row["训练目标"])
            cycle.end_date = parse_date(row["结束日期"])
            cycle.target_race_name = _optional_text(row["目标比赛名称"])
            cycle.target_race_date = parse_date(row["目标比赛日期"])
            cycle.target_result = _optional_text(row["目标成绩"])
            cycle.description = _optional_text(row["说明"])
            db.flush()
            stats["success"] += 1
            return cycle
        except ValueError as exc:
            _add_error(errors, SHEET_CYCLE, row_index, str(exc))
            return None
    _add_error(errors, SHEET_CYCLE, 2, "未找到有效训练周期。")
    return None


def _import_blocks(
    db: Session,
    sheet: Any,
    user_id: int,
    cycle: TrainingCycle | None,
    cycle_year: int | None,
    errors: list[ExcelImportErrorItem],
    stats: dict[str, int],
) -> dict[str, TrainingBlock]:
    blocks: dict[str, TrainingBlock] = {}
    if cycle is None:
        return blocks
    for row_index, row in _iter_data_rows(sheet, SHEET_HEADERS[SHEET_BLOCKS]):
        stats["total"] += 1
        try:
            block_name = _required_text(row["训练块名称"], "训练块名称")
            sort_order = _required_int(row["排序"], "排序")
            block_type = _parse_block_type(row["块类型"])
            block = db.scalar(
                select(TrainingBlock).where(
                    TrainingBlock.user_id == user_id,
                    TrainingBlock.cycle_id == cycle.id,
                    TrainingBlock.block_name == block_name,
                )
            )
            if block is None:
                conflict = db.scalar(
                    select(TrainingBlock).where(
                        TrainingBlock.user_id == user_id,
                        TrainingBlock.cycle_id == cycle.id,
                        TrainingBlock.sort_order == sort_order,
                    )
                )
                if conflict is not None:
                    raise ValueError("同一训练周期内排序已被其他训练块使用。")
                block = TrainingBlock(user_id=user_id, cycle_id=cycle.id, block_name=block_name)
                db.add(block)
            block.block_type = block_type
            block.week_index = _optional_int(row["周序号"])
            block.sort_order = sort_order
            block.date_range_text = _optional_text(row["日期范围"])
            block.target_distance_min_km = parse_decimal(row["目标跑量下限km"])
            block.target_distance_max_km = parse_decimal(row["目标跑量上限km"])
            block.planned_distance_km = parse_decimal(row["计划跑量km"])
            block.start_date = parse_date(row["开始日期"], default_year=cycle_year)
            block.end_date = parse_date(row["结束日期"], default_year=cycle_year)
            block.phase_name = _optional_text(row["阶段名称"])
            block.focus = _optional_text(row["训练重点"])
            db.flush()
            blocks[block_name] = block
            stats["success"] += 1
        except ValueError as exc:
            _add_error(errors, SHEET_BLOCKS, row_index, str(exc))
    return blocks


def _import_workouts(
    db: Session,
    sheet: Any,
    user_id: int,
    cycle: TrainingCycle | None,
    blocks: dict[str, TrainingBlock],
    cycle_year: int | None,
    errors: list[ExcelImportErrorItem],
    stats: dict[str, int],
) -> dict[Any, PlannedWorkout]:
    workouts: dict[Any, PlannedWorkout] = {}
    if cycle is None:
        return workouts
    for row_index, row in _iter_data_rows(sheet, SHEET_HEADERS[SHEET_WORKOUTS]):
        stats["total"] += 1
        try:
            workout_date = parse_date(row["日期"], default_year=cycle_year)
            if workout_date is None:
                raise ValueError("日期不能为空。")
            block_name = _required_text(row["训练块名称"], "训练块名称")
            block = blocks.get(block_name) or db.scalar(
                select(TrainingBlock).where(
                    TrainingBlock.user_id == user_id,
                    TrainingBlock.cycle_id == cycle.id,
                    TrainingBlock.block_name == block_name,
                )
            )
            if block is None:
                raise ValueError("找不到对应训练块。")
            planned_content = _required_text(row["计划训练内容"], "计划训练内容")
            raw_type = _optional_text(row["主类型"])
            workout = db.scalar(
                select(PlannedWorkout).where(
                    PlannedWorkout.user_id == user_id,
                    PlannedWorkout.cycle_id == cycle.id,
                    PlannedWorkout.workout_date == workout_date,
                )
            )
            if workout is None:
                workout = PlannedWorkout(
                    user_id=user_id,
                    cycle_id=cycle.id,
                    workout_date=workout_date,
                    workout_log=WorkoutLog(
                        user_id=user_id,
                        status_raw=None,
                        status_normalized=WorkoutStatusNormalized.not_started,
                    ),
                )
                db.add(workout)
            workout.block_id = block.id
            workout.date_text = workout_date.isoformat()
            workout.weekday = _optional_text(row["星期"])
            workout.phase_name = _optional_text(row["阶段名称"])
            workout.planned_content = planned_content
            workout.focus_note = _merge_focus_and_pace(row["重点说明"], row["目标配速"])
            workout.planned_distance_km = parse_decimal(row["计划km"])
            workout.main_type_raw = raw_type
            workout.main_type_normalized = normalize_workout_main_type(raw_type)
            workout.source_sheet = SHEET_WORKOUTS
            workout.source_row = row_index
            workout.sort_order = _required_int(row["排序"], "排序")
            db.flush()
            workouts[workout_date] = workout
            stats["success"] += 1
        except ValueError as exc:
            _add_error(errors, SHEET_WORKOUTS, row_index, str(exc))
    return workouts


def _import_logs(
    db: Session,
    sheet: Any,
    user_id: int,
    workouts: dict[Any, PlannedWorkout],
    cycle_year: int | None,
    errors: list[ExcelImportErrorItem],
    stats: dict[str, int],
) -> None:
    for row_index, row in _iter_data_rows(sheet, SHEET_HEADERS[SHEET_LOGS]):
        stats["total"] += 1
        try:
            workout_date = parse_date(row["日期"], default_year=cycle_year)
            if workout_date is None:
                raise ValueError("日期不能为空。")
            planned = workouts.get(workout_date)
            if planned is None:
                planned = db.scalar(
                    select(PlannedWorkout).where(
                        PlannedWorkout.user_id == user_id,
                        PlannedWorkout.workout_date == workout_date,
                    )
                )
            if planned is None:
                raise ValueError("找不到对应训练计划。")
            log = planned.workout_log or db.scalar(
                select(WorkoutLog).where(
                    WorkoutLog.user_id == user_id,
                    WorkoutLog.planned_workout_id == planned.id,
                )
            )
            if log is None:
                log = WorkoutLog(user_id=user_id, planned_workout_id=planned.id)
                db.add(log)
            pain_level = _optional_int(row["疼痛等级"])
            if pain_level is not None and not 0 <= pain_level <= 10:
                raise ValueError("疼痛等级必须在 0-10 之间。")
            status_raw = _optional_text(row["完成状态"])
            log.status_raw = status_raw
            log.status_normalized = normalize_workout_status(status_raw)
            log.actual_distance_km = parse_decimal(row["实际km"])
            log.actual_duration_seconds = parse_duration_seconds(row["实际时长"])
            log.avg_pace_seconds_per_km = parse_pace_seconds_per_km(row["均配"])
            log.avg_heart_rate = _optional_int(row["均心率"])
            log.rpe = _optional_int(row["RPE"])
            log.i_effective_km = parse_decimal(row["I有效km"])
            log.t1_effective_km = parse_decimal(row["T1有效km"])
            log.t2_effective_km = parse_decimal(row["T2有效km"])
            log.m_effective_km = parse_decimal(row["M有效km"])
            log.r_effective_km = parse_decimal(row["R短速度km"])
            log.sleep_hours = parse_decimal(row["睡眠h"])
            log.hrv = _optional_int(row["HRV"])
            log.morning_heart_rate = _optional_int(row["晨脉"])
            log.weight_kg = parse_decimal(row["体重kg"])
            log.leg_feeling = _optional_text(row["腿感"])
            log.pain_location = _optional_text(row["疼痛部位"])
            log.pain_level = pain_level
            log.pain_scale_version = PainScaleVersion.native_0_10
            log.main_session_data = _optional_text(row["主课数据"])
            log.review_note = _optional_text(row["一句复盘"])
            log.tomorrow_adjustment = _optional_text(row["明日调整"])
            log.alert_message = _optional_text(row["训练警报"])
            log.completion_rate = parse_decimal(row["达成率"])
            db.flush()
            stats["success"] += 1
        except ValueError as exc:
            _add_error(errors, SHEET_LOGS, row_index, str(exc))


def _import_reviews(
    db: Session,
    sheet: Any,
    user_id: int,
    blocks: dict[str, TrainingBlock],
    errors: list[ExcelImportErrorItem],
    stats: dict[str, int],
) -> None:
    for row_index, row in _iter_data_rows(sheet, SHEET_HEADERS[SHEET_REVIEWS]):
        stats["total"] += 1
        try:
            block_name = _required_text(row["训练块名称"], "训练块名称")
            block = blocks.get(block_name) or db.scalar(
                select(TrainingBlock).where(
                    TrainingBlock.user_id == user_id,
                    TrainingBlock.block_name == block_name,
                )
            )
            if block is None:
                raise ValueError("找不到训练块。")
            max_pain_level = _optional_int(row["最高疼痛等级"])
            if max_pain_level is not None and not 0 <= max_pain_level <= 10:
                raise ValueError("最高疼痛等级必须在 0-10 之间。")
            review = block.block_review or db.scalar(
                select(BlockReview).where(
                    BlockReview.user_id == user_id,
                    BlockReview.block_id == block.id,
                )
            )
            if review is None:
                review = BlockReview(user_id=user_id, block_id=block.id)
                db.add(review)
            review.planned_distance_km = parse_decimal(row["计划km"])
            review.actual_distance_km = parse_decimal(row["实际km"])
            review.completion_rate = parse_decimal(row["完成率"])
            review.i_effective_km = parse_decimal(row["I有效km"])
            review.t1_effective_km = parse_decimal(row["T1有效km"])
            review.t2_effective_km = parse_decimal(row["T2有效km"])
            review.m_effective_km = parse_decimal(row["M有效km"])
            review.r_effective_km = parse_decimal(row["R短速度km"])
            review.avg_rpe = parse_decimal(row["平均RPE"])
            review.avg_weight_kg = parse_decimal(row["平均体重kg"])
            review.max_pain_level = max_pain_level
            review.review_text = _optional_text(row["本周复盘"])
            review.next_block_adjustment = _optional_text(row["下周调整"])
            db.flush()
            stats["success"] += 1
        except ValueError as exc:
            _add_error(errors, SHEET_REVIEWS, row_index, str(exc))


def _import_pace_rules(
    db: Session,
    sheet: Any,
    user_id: int,
    errors: list[ExcelImportErrorItem],
    stats: dict[str, int],
) -> None:
    for row_index, row in _iter_data_rows(sheet, SHEET_HEADERS[SHEET_PACE_RULES]):
        stats["total"] += 1
        try:
            code = _required_text(row["代号"], "代号").upper()
            rule = db.scalar(select(PaceRule).where(PaceRule.user_id == user_id, PaceRule.code == code))
            if rule is None:
                rule = PaceRule(user_id=user_id, code=code)
                db.add(rule)
            rule.name = _required_text(row["类型"], "类型")
            rule.target_pace_text = _optional_text(row["目标配速"])
            rule.physiological_purpose = _optional_text(row["生理目的"])
            rule.note = _optional_text(row["备注"])
            rule.sort_order = _required_int(row["排序"], "排序")
            db.flush()
            stats["success"] += 1
        except ValueError as exc:
            _add_error(errors, SHEET_PACE_RULES, row_index, str(exc))


def _iter_data_rows(sheet: Any, headers: list[str]):
    for row_index in range(2, sheet.max_row + 1):
        values = [sheet.cell(row=row_index, column=index).value for index in range(1, len(headers) + 1)]
        if all(is_blank(value) for value in values):
            continue
        yield row_index, dict(zip(headers, values, strict=True))


def _parse_block_type(value: Any) -> BlockType:
    text = _required_text(value, "块类型")
    try:
        return BlockType(text)
    except ValueError as exc:
        raise ValueError("块类型必须是 week、transition 或 special。") from exc


def _required_text(value: Any, field_name: str) -> str:
    text = _optional_text(value)
    if text is None:
        raise ValueError(f"{field_name}不能为空。")
    return text


def _optional_text(value: Any) -> str | None:
    if is_blank(value):
        return None
    return str(value).strip()


def _required_int(value: Any, field_name: str) -> int:
    parsed = _optional_int(value)
    if parsed is None:
        raise ValueError(f"{field_name}不能为空。")
    return parsed


def _optional_int(value: Any) -> int | None:
    if is_blank(value):
        return None
    if isinstance(value, Decimal):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text.isdigit():
        raise ValueError("整数格式错误。")
    return int(text)


def _merge_focus_and_pace(focus: Any, target_pace: Any) -> str | None:
    focus_text = _optional_text(focus)
    pace_text = _optional_text(target_pace)
    if focus_text and pace_text:
        return f"{focus_text}\n目标配速：{pace_text}"
    if pace_text:
        return f"目标配速：{pace_text}"
    return focus_text


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _add_error(errors: list[ExcelImportErrorItem], sheet: str, row: int, message: str) -> None:
    errors.append(ExcelImportErrorItem(sheet=sheet, row=row, message=message))


def _result_status(success_count: int, failed_count: int) -> str:
    if success_count == 0 and failed_count > 0:
        return "failed"
    if failed_count > 0:
        return "partial_success"
    return "success"


def _job_status(status: str) -> ExcelImportStatus:
    if status == "partial_success":
        return ExcelImportStatus.partial_success
    if status == "failed":
        return ExcelImportStatus.failed
    return ExcelImportStatus.success


def _failed_template_result(message: str) -> ExcelImportResult:
    return ExcelImportResult(
        status="failed",
        message="Excel 模板格式错误",
        total_count=0,
        success_count=0,
        failed_count=0,
        errors=[ExcelImportErrorItem(sheet="Excel", row=0, message=message)],
    )
