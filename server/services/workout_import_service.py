from __future__ import annotations

import csv
import io
import json
import re
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from planner_core.database.models import (
    PlannedWorkout,
    WorkoutImportAudit,
    WorkoutImportBatch,
    WorkoutImportItem,
    WorkoutLog,
)
from planner_core.enums import PainScaleVersion, WorkoutStatusNormalized
from planner_core.utils.excel_parse import (
    is_blank,
    parse_date,
    parse_decimal,
    parse_duration_seconds,
    parse_pace_seconds_per_km,
)
from server.common.exceptions import BadRequestError, NotFoundError
from server.schemas.workout_import import (
    NormalizedWorkoutActivity,
    WorkoutImportApplyResponse,
    WorkoutImportBatchRead,
    WorkoutImportCreateResponse,
    WorkoutImportFieldDiff,
    WorkoutImportIssue,
    WorkoutImportItemPatch,
    WorkoutImportItemRead,
    WorkoutImportPreviewSummary,
    WorkoutImportStructuredRequest,
)
from server.services.import_core import bytes_hash, payload_hash, validate_upload_basics

PARSER_VERSION = "workout-import-parser-v1"
NORMALIZATION_VERSION = "workout-activity-v1"
MAX_FILE_SIZE_BYTES = 2 * 1024 * 1024
MAX_ROWS = 1000
SUPPORTED_EXTENSIONS = {".json", ".xlsx", ".csv", ".txt", ".md"}
ALLOWED_MIME_TYPES = {
    ".json": {"application/json", "text/plain", "application/octet-stream"},
    ".xlsx": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/octet-stream"},
    ".csv": {"text/csv", "text/plain", "application/vnd.ms-excel", "application/octet-stream"},
    ".txt": {"text/plain", "application/octet-stream"},
    ".md": {"text/markdown", "text/plain", "application/octet-stream"},
}

OBJECTIVE_FIELD_MAP = {
    "distance_km": "actual_distance_km",
    "duration_seconds": "actual_duration_seconds",
    "moving_time_seconds": "moving_time_seconds",
    "elapsed_time_seconds": "elapsed_time_seconds",
    "average_pace_seconds_per_km": "avg_pace_seconds_per_km",
    "average_heart_rate_bpm": "avg_heart_rate",
    "max_heart_rate_bpm": "max_heart_rate",
    "average_cadence_spm": "average_cadence_spm",
    "max_cadence_spm": "max_cadence_spm",
    "elevation_gain_m": "elevation_gain_m",
    "calories_kcal": "calories_kcal",
}
SUBJECTIVE_FIELD_MAP = {
    "rpe": "rpe",
    "pain_level": "pain_level",
    "notes": "review_note",
    "content": "main_session_data",
}
TEXT_HEADERS = ["activity_date", "session_index", "workout_type", "distance_km", "duration_seconds", "average_heart_rate_bpm", "rpe", "title", "notes"]


def create_structured_import(
    db: Session,
    user_id: int,
    payload: WorkoutImportStructuredRequest,
    idempotency_key: str | None = None,
) -> WorkoutImportCreateResponse:
    request_id = idempotency_key or payload.client_request_id
    existing = _find_existing_batch(db, user_id, request_id)
    if existing:
        return _create_response(existing)
    raw_hash = payload_hash(payload.model_dump(mode="json"))
    return _create_batch(db, user_id, payload, source_type=payload.source or "structured_json", source_filename=None, raw_payload_hash=raw_hash)


async def create_file_import(
    db: Session,
    user_id: int,
    *,
    filename: str,
    content_type: str | None,
    content: bytes,
    merge_strategy: str,
    timezone: str,
    client_request_id: str | None,
    idempotency_key: str | None = None,
) -> WorkoutImportCreateResponse:
    clean_name, extension = validate_upload_basics(
        filename=filename,
        content_type=content_type,
        content=content,
        supported_extensions=SUPPORTED_EXTENSIONS,
        allowed_mime_types=ALLOWED_MIME_TYPES,
        max_file_size_bytes=MAX_FILE_SIZE_BYTES,
    )
    request_id = idempotency_key or client_request_id
    existing = _find_existing_batch(db, user_id, request_id)
    if existing:
        return _create_response(existing)
    activities = _parse_file(extension, content)
    payload = WorkoutImportStructuredRequest(
        source=_source_type_for_extension(extension),
        timezone=timezone,
        merge_strategy=merge_strategy,  # type: ignore[arg-type]
        client_request_id=client_request_id,
        activities=activities,
    )
    return _create_batch(db, user_id, payload, source_type=_source_type_for_extension(extension), source_filename=clean_name, raw_payload_hash=bytes_hash(content))


def list_workout_imports(db: Session, user_id: int, limit: int = 20) -> list[WorkoutImportBatchRead]:
    batches = db.scalars(
        select(WorkoutImportBatch)
        .options(selectinload(WorkoutImportBatch.items))
        .where(WorkoutImportBatch.user_id == user_id)
        .order_by(WorkoutImportBatch.created_at.desc())
        .limit(min(limit, 100))
    ).all()
    return [WorkoutImportBatchRead.model_validate(batch) for batch in batches]


def get_workout_import(db: Session, user_id: int, batch_id: int) -> WorkoutImportBatchRead:
    return WorkoutImportBatchRead.model_validate(_get_batch(db, user_id, batch_id))


def update_workout_import_item(
    db: Session,
    user_id: int,
    batch_id: int,
    item_id: int,
    payload: WorkoutImportItemPatch,
) -> WorkoutImportBatchRead:
    batch = _get_batch(db, user_id, batch_id)
    _ensure_editable(batch)
    item = next((candidate for candidate in batch.items if candidate.id == item_id), None)
    if item is None:
        raise NotFoundError("导入条目不存在。")
    if payload.matched_plan_id is not None:
        plan = _get_user_plan(db, user_id, payload.matched_plan_id)
        item.matched_plan_id = plan.id
    if payload.normalized_data is not None:
        item.normalized_data_json = payload.normalized_data.model_dump(mode="json")
        item.activity_date = payload.normalized_data.activity_date
        item.start_time = payload.normalized_data.start_time
        item.session_index = payload.normalized_data.session_index
    if payload.session_index is not None:
        data = dict(item.normalized_data_json or {})
        data["session_index"] = payload.session_index
        item.normalized_data_json = NormalizedWorkoutActivity.model_validate(data).model_dump(mode="json")
        item.session_index = payload.session_index
    if payload.workout_type is not None:
        data = dict(item.normalized_data_json or {})
        data["workout_type"] = payload.workout_type
        item.normalized_data_json = NormalizedWorkoutActivity.model_validate(data).model_dump(mode="json")
    if payload.user_action is not None:
        item.user_action = payload.user_action
    db.commit()
    return validate_workout_import(db, user_id, batch_id)


def validate_workout_import(db: Session, user_id: int, batch_id: int) -> WorkoutImportBatchRead:
    batch = _get_batch(db, user_id, batch_id)
    _ensure_editable(batch)
    activities = [NormalizedWorkoutActivity.model_validate(item.normalized_data_json) for item in batch.items if item.normalized_data_json]
    preview = _build_preview(db, user_id, activities, batch.merge_strategy, batch.source_type)
    _replace_items(batch, preview, preserve_actions={item.activity_fingerprint: item.user_action for item in batch.items if item.activity_fingerprint and item.user_action})
    _apply_summary_to_batch(batch, preview)
    db.commit()
    return WorkoutImportBatchRead.model_validate(_get_batch(db, user_id, batch_id))


def apply_workout_import(db: Session, user_id: int, batch_id: int) -> WorkoutImportApplyResponse:
    try:
        batch = _get_batch(db, user_id, batch_id, lock=True)
        if batch.status == "applied":
            audit = batch.audit
            return WorkoutImportApplyResponse(
                batch_id=batch.id,
                status=batch.status,
                created_count=audit.created_count if audit else 0,
                updated_count=audit.updated_count if audit else 0,
                linked_plan_count=audit.linked_plan_count if audit else 0,
                unplanned_count=audit.unplanned_count if audit else 0,
                skipped_count=audit.skipped_count if audit else 0,
            )
        if batch.status in {"cancelled", "expired"}:
            raise BadRequestError("该导入草稿已不可应用。")
        if batch.expires_at and batch.expires_at < datetime.utcnow():
            batch.status = "expired"
            db.commit()
            raise BadRequestError("该导入草稿已过期。")

        activities = [NormalizedWorkoutActivity.model_validate(item.normalized_data_json) for item in batch.items if item.normalized_data_json]
        preview = _build_preview(db, user_id, activities, batch.merge_strategy, batch.source_type)
        preserved_actions = {item.activity_fingerprint: item.user_action for item in batch.items if item.activity_fingerprint and item.user_action}
        _replace_items(batch, preview, preserve_actions=preserved_actions)
        _apply_summary_to_batch(batch, preview)
        if batch.conflict_count or batch.invalid_count:
            db.commit()
            raise BadRequestError("导入草稿仍存在冲突或无效数据，不能应用。")

        created_count = updated_count = linked_plan_count = unplanned_count = skipped_count = subjective_missing_count = 0
        db.flush()
        for item in batch.items:
            activity = NormalizedWorkoutActivity.model_validate(item.normalized_data_json)
            action = item.user_action or item.suggested_action
            applied_log: WorkoutLog | None = None
            if action in {"skip", "keep_existing", "manual_review"}:
                skipped_count += 1
                continue
            log = item.matched_log
            if action in {"fill_empty_fields", "update_objective_fields"} and log is not None:
                _merge_into_log(log, activity, action, batch.source_type, batch.id)
                updated_count += 1
                item.applied_log_id = log.id
                applied_log = log
            elif action in {"create_log", "create_unplanned_log", "link_to_plan"}:
                plan = item.matched_plan
                if plan is not None and plan.workout_log is not None:
                    skipped_count += 1
                    continue
                log = _new_log(user_id, activity, batch.source_type, batch.id, plan)
                db.add(log)
                db.flush()
                item.applied_log_id = log.id
                applied_log = log
                created_count += 1
                if plan is not None:
                    linked_plan_count += 1
                else:
                    unplanned_count += 1
            if applied_log and _missing_subjective(applied_log):
                subjective_missing_count += 1

        batch.status = "applied"
        batch.applied_at = datetime.utcnow()
        db.add(
            WorkoutImportAudit(
                user_id=user_id,
                batch_id=batch.id,
                source_type=batch.source_type,
                merge_strategy=batch.merge_strategy,
                total_count=batch.total_count,
                created_count=created_count,
                updated_count=updated_count,
                linked_plan_count=linked_plan_count,
                unplanned_count=unplanned_count,
                skipped_count=skipped_count,
                conflict_count=batch.conflict_count,
                applied_at=batch.applied_at,
                actor_type="user",
                client_request_id=batch.client_request_id,
            )
        )
        db.commit()
        return WorkoutImportApplyResponse(
            batch_id=batch.id,
            status=batch.status,
            created_count=created_count,
            updated_count=updated_count,
            linked_plan_count=linked_plan_count,
            unplanned_count=unplanned_count,
            skipped_count=skipped_count,
            subjective_missing_count=subjective_missing_count,
        )
    except Exception:
        db.rollback()
        raise


def cancel_workout_import(db: Session, user_id: int, batch_id: int) -> WorkoutImportBatchRead:
    batch = _get_batch(db, user_id, batch_id)
    if batch.status == "applied":
        raise BadRequestError("已应用的导入草稿不能取消。")
    batch.status = "cancelled"
    batch.cancelled_at = datetime.utcnow()
    db.commit()
    return WorkoutImportBatchRead.model_validate(batch)


def _find_existing_batch(db: Session, user_id: int, request_id: str | None) -> WorkoutImportBatch | None:
    if not request_id:
        return None
    return db.scalar(
        select(WorkoutImportBatch)
        .options(selectinload(WorkoutImportBatch.items))
        .where(WorkoutImportBatch.user_id == user_id, WorkoutImportBatch.client_request_id == request_id)
    )


def _get_batch(db: Session, user_id: int, batch_id: int, *, lock: bool = False) -> WorkoutImportBatch:
    stmt = (
        select(WorkoutImportBatch)
        .options(
            selectinload(WorkoutImportBatch.items).selectinload(WorkoutImportItem.matched_plan),
            selectinload(WorkoutImportBatch.items).selectinload(WorkoutImportItem.matched_log),
            selectinload(WorkoutImportBatch.items).selectinload(WorkoutImportItem.applied_log),
            selectinload(WorkoutImportBatch.audit),
        )
        .where(WorkoutImportBatch.id == batch_id, WorkoutImportBatch.user_id == user_id)
    )
    if lock:
        stmt = stmt.with_for_update()
    batch = db.scalar(stmt)
    if batch is None:
        raise NotFoundError("导入草稿不存在。")
    batch.items.sort(key=lambda item: (item.activity_date or date.max, item.session_index or 99, item.id))
    return batch


def _create_batch(
    db: Session,
    user_id: int,
    payload: WorkoutImportStructuredRequest,
    *,
    source_type: str,
    source_filename: str | None,
    raw_payload_hash: str,
) -> WorkoutImportCreateResponse:
    preview = _build_preview(db, user_id, payload.activities[:MAX_ROWS], payload.merge_strategy, source_type)
    batch = WorkoutImportBatch(
        user_id=user_id,
        source_type=source_type,
        source_filename=source_filename,
        parser_version=PARSER_VERSION,
        normalization_version=NORMALIZATION_VERSION,
        raw_payload_hash=raw_payload_hash,
        merge_strategy=payload.merge_strategy,
        status="conflict" if preview["summary"].conflict_count else ("validation_failed" if preview["summary"].invalid_count else "ready"),
        timezone=payload.timezone,
        client_request_id=payload.client_request_id,
        expires_at=datetime.utcnow() + timedelta(days=14),
    )
    _replace_items(batch, preview, preserve_actions={})
    _apply_summary_to_batch(batch, preview)
    db.add(batch)
    db.commit()
    db.refresh(batch)
    return _create_response(_get_batch(db, user_id, batch.id))


def _create_response(batch: WorkoutImportBatch) -> WorkoutImportCreateResponse:
    summary = WorkoutImportPreviewSummary.model_validate(batch.preview_summary_json or {})
    warnings = [WorkoutImportIssue.model_validate(item) for item in batch.warnings_json or []]
    return WorkoutImportCreateResponse(
        batch_id=batch.id,
        status=batch.status,
        total_count=batch.total_count,
        matched_plan_count=batch.matched_plan_count,
        matched_log_count=batch.matched_log_count,
        unplanned_count=batch.unplanned_count,
        ready_count=batch.ready_count,
        conflict_count=batch.conflict_count,
        invalid_count=batch.invalid_count,
        skipped_count=batch.skipped_count,
        warnings=warnings,
        items=[WorkoutImportItemRead.model_validate(item) for item in batch.items],
        preview_summary=summary,
    )


def _build_preview(db: Session, user_id: int, activities: list[NormalizedWorkoutActivity], merge_strategy: str, source_type: str) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    summary = WorkoutImportPreviewSummary(total_count=len(activities))
    batch_warnings: list[WorkoutImportIssue] = []
    for row_number, activity in enumerate(activities, start=1):
        errors: list[WorkoutImportIssue] = []
        warnings: list[WorkoutImportIssue] = []
        if activity.activity_date > date.today() + timedelta(days=1):
            errors.append(WorkoutImportIssue(code="FUTURE_ACTIVITY", message="已完成训练不能是明显未来日期。", row_number=row_number, field="activity_date"))
        fingerprint = _activity_fingerprint(source_type, activity)
        plan, plan_status, confidence = _match_plan(db, user_id, activity)
        log = _match_log(db, user_id, activity, fingerprint, plan, source_type)
        field_diff = _field_diff(log, activity, source_type) if log else []
        if errors:
            match_status = "invalid"
            suggested_action = "skip"
            summary.invalid_count += 1
        elif plan_status == "ambiguous":
            match_status = "ambiguous"
            suggested_action = "manual_review"
            summary.conflict_count += 1
        elif log:
            match_status = "matched_existing_log"
            summary.matched_log_count += 1
            suggested_action = _existing_log_action(log, activity, merge_strategy)
            if suggested_action in {"skip", "keep_existing"}:
                summary.skipped_count += 1
            else:
                summary.ready_count += 1
        elif plan:
            match_status = "matched_plan"
            summary.matched_plan_count += 1
            suggested_action = "manual_review" if merge_strategy == "manual_review" else "create_log"
            summary.ready_count += 1 if suggested_action != "manual_review" else 0
            summary.conflict_count += 1 if suggested_action == "manual_review" else 0
        else:
            match_status = "unplanned_activity"
            suggested_action = "manual_review" if merge_strategy == "manual_review" else "create_unplanned_log"
            summary.unplanned_count += 1
            summary.ready_count += 1 if suggested_action != "manual_review" else 0
            summary.conflict_count += 1 if suggested_action == "manual_review" else 0
        if warnings:
            batch_warnings.extend(warnings)
        rows.append(
            {
                "row_number": row_number,
                "activity": activity,
                "plan": plan,
                "log": log,
                "match_status": match_status,
                "match_confidence": confidence,
                "suggested_action": suggested_action,
                "errors": errors,
                "warnings": warnings,
                "field_diff": field_diff,
                "fingerprint": fingerprint,
            }
        )
    return {"rows": rows, "summary": summary, "warnings": batch_warnings}


def _replace_items(batch: WorkoutImportBatch, preview: dict[str, Any], preserve_actions: dict[str, str | None]) -> None:
    batch.items.clear()
    for row in preview["rows"]:
        activity = row["activity"]
        fingerprint = row["fingerprint"]
        batch.items.append(
            WorkoutImportItem(
                row_number=row["row_number"],
                activity_date=activity.activity_date,
                start_time=activity.start_time,
                session_index=activity.session_index,
                normalized_data_json=activity.model_dump(mode="json"),
                matched_plan_id=row["plan"].id if row["plan"] else None,
                matched_log_id=row["log"].id if row["log"] else None,
                match_status=row["match_status"],
                match_confidence=row["match_confidence"],
                suggested_action=row["suggested_action"],
                user_action=preserve_actions.get(fingerprint),
                validation_errors_json=[issue.model_dump(mode="json") for issue in row["errors"]],
                warnings_json=[issue.model_dump(mode="json") for issue in row["warnings"]],
                field_diff_json=[diff.model_dump(mode="json") for diff in row["field_diff"]],
                activity_fingerprint=fingerprint,
            )
        )


def _apply_summary_to_batch(batch: WorkoutImportBatch, preview: dict[str, Any]) -> None:
    summary = preview["summary"]
    batch.total_count = summary.total_count
    batch.matched_plan_count = summary.matched_plan_count
    batch.matched_log_count = summary.matched_log_count
    batch.unplanned_count = summary.unplanned_count
    batch.ready_count = summary.ready_count
    batch.conflict_count = summary.conflict_count
    batch.invalid_count = summary.invalid_count
    batch.skipped_count = summary.skipped_count
    batch.warnings_json = [issue.model_dump(mode="json") for issue in preview["warnings"]]
    batch.preview_summary_json = summary.model_dump(mode="json")
    batch.status = "conflict" if summary.conflict_count else ("validation_failed" if summary.invalid_count else "ready")


def _match_plan(db: Session, user_id: int, activity: NormalizedWorkoutActivity) -> tuple[PlannedWorkout | None, str, str | None]:
    if activity.planned_workout_id:
        plan = _get_user_plan(db, user_id, activity.planned_workout_id)
        return plan, "matched_plan", "high"
    candidates = list(
        db.scalars(
            select(PlannedWorkout)
            .options(selectinload(PlannedWorkout.workout_log))
            .where(
                PlannedWorkout.user_id == user_id,
                PlannedWorkout.workout_date == activity.activity_date,
                PlannedWorkout.session_index == activity.session_index,
            )
        )
    )
    if len(candidates) == 1:
        return candidates[0], "matched_plan", "high"
    if len(candidates) > 1:
        return None, "ambiguous", "low"
    same_day = list(
        db.scalars(
            select(PlannedWorkout)
            .options(selectinload(PlannedWorkout.workout_log))
            .where(PlannedWorkout.user_id == user_id, PlannedWorkout.workout_date == activity.activity_date)
        )
    )
    typed = [plan for plan in same_day if activity.workout_type and (plan.main_type_raw or "").upper() == activity.workout_type.upper()]
    if len(typed) == 1:
        return typed[0], "matched_plan", "medium"
    return None, "unplanned_activity", None


def _match_log(db: Session, user_id: int, activity: NormalizedWorkoutActivity, fingerprint: str, plan: PlannedWorkout | None, source_type: str) -> WorkoutLog | None:
    if activity.external_activity_id:
        log = db.scalar(
            select(WorkoutLog).where(
                WorkoutLog.user_id == user_id,
                WorkoutLog.external_activity_id == activity.external_activity_id,
                WorkoutLog.source_type == source_type,
            )
        )
        if log:
            return log
    log = db.scalar(select(WorkoutLog).where(WorkoutLog.user_id == user_id, WorkoutLog.activity_fingerprint == fingerprint))
    if log:
        return log
    if plan is not None and plan.workout_log is not None:
        return plan.workout_log
    return db.scalar(
        select(WorkoutLog).where(
            WorkoutLog.user_id == user_id,
            WorkoutLog.activity_date == activity.activity_date,
            WorkoutLog.session_index == activity.session_index,
            WorkoutLog.start_time == activity.start_time,
        )
    )


def _existing_log_action(log: WorkoutLog, activity: NormalizedWorkoutActivity, merge_strategy: str) -> str:
    if merge_strategy == "manual_review":
        return "manual_review"
    if merge_strategy == "fill_empty_fields" and _fillable_objective_fields(log, activity):
        return "fill_empty_fields"
    if merge_strategy == "update_objective_fields" and _incoming_objective_fields(activity):
        return "update_objective_fields"
    return "keep_existing"


def _field_diff(log: WorkoutLog, activity: NormalizedWorkoutActivity, source_type: str) -> list[WorkoutImportFieldDiff]:
    diffs: list[WorkoutImportFieldDiff] = []
    sources = log.field_sources_json or {}
    for incoming_field, log_field in {**OBJECTIVE_FIELD_MAP, **SUBJECTIVE_FIELD_MAP}.items():
        incoming = getattr(activity, incoming_field)
        existing = getattr(log, log_field)
        if incoming != existing and incoming is not None:
            diffs.append(
                WorkoutImportFieldDiff(
                    field=incoming_field,
                    existing_value=existing,
                    incoming_value=incoming,
                    existing_source=sources.get(log_field) or log.source_type,
                    incoming_source=source_type,
                    recommended_action=("keep_existing" if incoming_field in SUBJECTIVE_FIELD_MAP else "strategy_dependent"),
                )
            )
    return diffs


def _merge_into_log(log: WorkoutLog, activity: NormalizedWorkoutActivity, action: str, source_type: str, batch_id: int) -> None:
    if action == "fill_empty_fields":
        for incoming_field, log_field in OBJECTIVE_FIELD_MAP.items():
            value = getattr(activity, incoming_field)
            if getattr(log, log_field) is None and value is not None:
                setattr(log, log_field, value)
                _mark_field_source(log, log_field, source_type)
    elif action == "update_objective_fields":
        for incoming_field, log_field in OBJECTIVE_FIELD_MAP.items():
            value = getattr(activity, incoming_field)
            if value is not None:
                setattr(log, log_field, value)
                _mark_field_source(log, log_field, source_type)
    _fill_identity_fields(log, activity, source_type, batch_id)


def _new_log(user_id: int, activity: NormalizedWorkoutActivity, source_type: str, batch_id: int, plan: PlannedWorkout | None) -> WorkoutLog:
    log = WorkoutLog(
        user_id=user_id,
        planned_workout_id=plan.id if plan else None,
        cycle_id=plan.cycle_id if plan else None,
        cycle_assignment_status="assigned" if plan else "unassigned",
        status_raw="completed",
        status_normalized=WorkoutStatusNormalized.completed_normal,
        pain_scale_version=PainScaleVersion.native_0_10,
        activity_date=activity.activity_date,
        start_time=activity.start_time,
        timezone=activity.timezone,
        session_index=activity.session_index,
        sport_type=activity.sport_type,
        workout_type=activity.workout_type,
        title=activity.title,
        is_unplanned=plan is None,
        source_type=source_type,
        source_import_batch_id=batch_id,
        external_activity_id=activity.external_activity_id,
        activity_fingerprint=_activity_fingerprint(source_type, activity),
        field_sources_json={},
    )
    for incoming_field, log_field in OBJECTIVE_FIELD_MAP.items():
        value = getattr(activity, incoming_field)
        if value is not None:
            setattr(log, log_field, value)
            _mark_field_source(log, log_field, source_type)
    for incoming_field, log_field in SUBJECTIVE_FIELD_MAP.items():
        value = getattr(activity, incoming_field)
        if value is not None:
            setattr(log, log_field, value)
            _mark_field_source(log, log_field, source_type)
    return log


def _fill_identity_fields(log: WorkoutLog, activity: NormalizedWorkoutActivity, source_type: str, batch_id: int) -> None:
    log.activity_date = log.activity_date or activity.activity_date
    log.start_time = log.start_time or activity.start_time
    log.timezone = log.timezone or activity.timezone
    log.session_index = log.session_index or activity.session_index
    log.sport_type = log.sport_type or activity.sport_type
    log.workout_type = log.workout_type or activity.workout_type
    log.title = log.title or activity.title
    log.source_import_batch_id = log.source_import_batch_id or batch_id
    log.activity_fingerprint = log.activity_fingerprint or _activity_fingerprint(source_type, activity)
    log.external_activity_id = log.external_activity_id or activity.external_activity_id


def _mark_field_source(log: WorkoutLog, field: str, source_type: str) -> None:
    sources = dict(log.field_sources_json or {})
    sources[field] = source_type
    log.field_sources_json = sources


def _fillable_objective_fields(log: WorkoutLog, activity: NormalizedWorkoutActivity) -> list[str]:
    return [log_field for incoming_field, log_field in OBJECTIVE_FIELD_MAP.items() if getattr(log, log_field) is None and getattr(activity, incoming_field) is not None]


def _incoming_objective_fields(activity: NormalizedWorkoutActivity) -> list[str]:
    return [log_field for incoming_field, log_field in OBJECTIVE_FIELD_MAP.items() if getattr(activity, incoming_field) is not None]


def _missing_subjective(log: WorkoutLog) -> bool:
    return log.rpe is None or log.leg_feeling is None or log.pain_level is None or log.review_note is None


def _activity_fingerprint(source_type: str, activity: NormalizedWorkoutActivity) -> str:
    if activity.external_activity_id:
        base = {"source": source_type, "external_activity_id": activity.external_activity_id}
    else:
        base = {
            "activity_date": activity.activity_date.isoformat(),
            "start_time": activity.start_time.isoformat() if activity.start_time else None,
            "session_index": activity.session_index,
            "duration_seconds": activity.duration_seconds,
            "distance_km": str(activity.distance_km) if activity.distance_km is not None else None,
            "workout_type": activity.workout_type,
        }
    return payload_hash(base)


def _get_user_plan(db: Session, user_id: int, plan_id: int) -> PlannedWorkout:
    plan = db.scalar(
        select(PlannedWorkout)
        .options(selectinload(PlannedWorkout.workout_log))
        .where(PlannedWorkout.id == plan_id, PlannedWorkout.user_id == user_id)
    )
    if plan is None:
        raise NotFoundError("训练计划不存在或不属于当前用户。")
    return plan


def _ensure_editable(batch: WorkoutImportBatch) -> None:
    if batch.status in {"applied", "cancelled", "expired"}:
        raise BadRequestError("该导入草稿已不可编辑。")


def _parse_file(extension: str, content: bytes) -> list[NormalizedWorkoutActivity]:
    if extension == ".json":
        try:
            body = json.loads(content.decode("utf-8-sig"))
        except json.JSONDecodeError as exc:
            raise BadRequestError(f"JSON 格式错误：第 {exc.lineno} 行。") from exc
        rows = body.get("activities", body) if isinstance(body, dict) else body
        if not isinstance(rows, list):
            raise BadRequestError("JSON 文件必须包含 activities 数组。")
        return _validate_rows(rows, "json")
    if extension == ".xlsx":
        return _parse_xlsx(content)
    if extension == ".csv":
        return _parse_csv(content.decode("utf-8-sig").splitlines())
    return _parse_text(content.decode("utf-8-sig").splitlines())


def _parse_xlsx(content: bytes) -> list[NormalizedWorkoutActivity]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    if getattr(workbook, "vba_archive", None) is not None:
        raise BadRequestError("不支持带宏 Excel 文件。")
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if row_number > MAX_ROWS + 1:
            raise BadRequestError("导入行数超过限制。")
        values = [cell.value for cell in row]
        if not any(str(value or "").strip() for value in values):
            continue
        for cell in row:
            if cell.data_type == "f":
                raise BadRequestError(f"Excel 第 {row_number} 行包含公式，已拒绝导入。")
        rows.append(dict(zip(headers, values, strict=False)))
    return _rows_to_activities(rows, "xlsx")


def _parse_csv(lines: list[str]) -> list[NormalizedWorkoutActivity]:
    return _rows_to_activities(list(csv.DictReader(lines)), "csv")


def _parse_text(lines: list[str]) -> list[NormalizedWorkoutActivity]:
    rows: list[dict[str, Any]] = []
    for row_number, raw in enumerate(lines, start=1):
        line = raw.strip().strip("|")
        if not line or line.startswith("#"):
            continue
        parts = [part.strip() for part in line.split("|")]
        if parts and parts[0] in {"日期", "activity_date"}:
            continue
        if len(parts) < 5:
            raise BadRequestError(f"文本第 {row_number} 行格式错误。")
        row = dict(zip(TEXT_HEADERS, parts, strict=False))
        rows.append(row)
    return _rows_to_activities(rows, "text")


def _rows_to_activities(rows: list[dict[str, Any]], source: str) -> list[NormalizedWorkoutActivity]:
    if not rows:
        raise BadRequestError("文件中没有可导入的训练记录。")
    activities: list[NormalizedWorkoutActivity] = []
    for row_number, row in enumerate(rows, start=2):
        if row_number > MAX_ROWS + 1:
            raise BadRequestError("导入行数超过限制。")
        try:
            activity_date = parse_date(row.get("activity_date") or row.get("date"))
            if activity_date is None:
                raise ValueError("activity_date 必填。")
            activities.append(
                NormalizedWorkoutActivity(
                    activity_date=activity_date,
                    start_time=_parse_time(row.get("start_time")),
                    timezone=_text_or_none(row.get("timezone")),
                    session_index=_int_or_none(row.get("session_index")) or 1,
                    sport_type=_text_or_none(row.get("sport_type")) or "running",
                    workout_type=_text_or_none(row.get("workout_type")),
                    title=_text_or_none(row.get("title")),
                    planned_workout_id=_int_or_none(row.get("planned_workout_id")),
                    distance_km=parse_decimal(row.get("distance_km")),
                    duration_seconds=parse_duration_seconds(row.get("duration_seconds")),
                    moving_time_seconds=parse_duration_seconds(row.get("moving_time_seconds")),
                    elapsed_time_seconds=parse_duration_seconds(row.get("elapsed_time_seconds")),
                    average_pace_seconds_per_km=parse_pace_seconds_per_km(row.get("average_pace_seconds_per_km")),
                    average_heart_rate_bpm=_int_or_none(row.get("average_heart_rate_bpm")),
                    max_heart_rate_bpm=_int_or_none(row.get("max_heart_rate_bpm")),
                    average_cadence_spm=_int_or_none(row.get("average_cadence_spm")),
                    max_cadence_spm=_int_or_none(row.get("max_cadence_spm")),
                    elevation_gain_m=_int_or_none(row.get("elevation_gain_m")),
                    calories_kcal=_int_or_none(row.get("calories_kcal")),
                    rpe=_int_or_none(row.get("rpe")),
                    pain_level=_int_or_none(row.get("pain_level")),
                    completion_status=_text_or_none(row.get("completion_status")) or "completed",
                    content=_text_or_none(row.get("content")),
                    notes=_text_or_none(row.get("notes")),
                    external_activity_id=_text_or_none(row.get("external_activity_id")),
                    source=source,
                )
            )
        except (ValidationError, ValueError) as exc:
            raise BadRequestError(f"{source} 第 {row_number} 行格式错误：{exc}") from exc
    return activities


def _validate_rows(rows: list[Any], source: str) -> list[NormalizedWorkoutActivity]:
    activities: list[NormalizedWorkoutActivity] = []
    for row_number, row in enumerate(rows, start=1):
        try:
            activities.append(NormalizedWorkoutActivity.model_validate(row))
        except ValidationError as exc:
            raise BadRequestError(f"{source} 第 {row_number} 行格式错误：{exc}") from exc
    if not activities:
        raise BadRequestError("文件中没有可导入的训练记录。")
    return activities


def _parse_time(value: Any) -> time | None:
    if is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    text = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            pass
    if re.fullmatch(r"\d+(\.\d+)?", text):
        seconds = int(float(text) * 86400) if float(text) < 1 else int(float(text))
        seconds %= 86400
        return time(seconds // 3600, (seconds % 3600) // 60, seconds % 60)
    raise ValueError("时间格式错误。")


def _source_type_for_extension(extension: str) -> str:
    return {".json": "json_import", ".xlsx": "excel_import", ".csv": "csv_import", ".txt": "text_import", ".md": "text_import"}[extension]


def _text_or_none(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None


def _int_or_none(value: Any) -> int | None:
    if is_blank(value):
        return None
    return int(value)
