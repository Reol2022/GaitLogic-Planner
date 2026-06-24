from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

TEMPLATE_FILENAME = "gaitlogic_planner_template.xlsx"

SHEET_INSTRUCTIONS = "填写说明"
SHEET_CYCLE = "训练周期"
SHEET_BLOCKS = "训练块"
SHEET_WORKOUTS = "训练计划"
SHEET_LOGS = "训练日志"
SHEET_REVIEWS = "每周复盘"
SHEET_PACE_RULES = "配速规则"

STANDARD_SHEETS = [
    SHEET_INSTRUCTIONS,
    SHEET_CYCLE,
    SHEET_BLOCKS,
    SHEET_WORKOUTS,
    SHEET_LOGS,
    SHEET_REVIEWS,
    SHEET_PACE_RULES,
]

SHEET_HEADERS: dict[str, list[str]] = {
    SHEET_CYCLE: [
        "周期名称",
        "训练目标",
        "开始日期",
        "结束日期",
        "目标比赛名称",
        "目标比赛日期",
        "目标成绩",
        "说明",
    ],
    SHEET_BLOCKS: [
        "训练块名称",
        "块类型",
        "周序号",
        "排序",
        "日期范围",
        "目标跑量下限km",
        "目标跑量上限km",
        "计划跑量km",
        "开始日期",
        "结束日期",
        "阶段名称",
        "训练重点",
    ],
    SHEET_WORKOUTS: [
        "日期",
        "星期",
        "训练块名称",
        "阶段名称",
        "计划训练内容",
        "重点说明",
        "计划km",
        "主类型",
        "目标配速",
        "排序",
    ],
    SHEET_LOGS: [
        "日期",
        "完成状态",
        "实际km",
        "实际时长",
        "均配",
        "均心率",
        "RPE",
        "I有效km",
        "T1有效km",
        "T2有效km",
        "M有效km",
        "R短速度km",
        "睡眠h",
        "HRV",
        "晨脉",
        "体重kg",
        "腿感",
        "疼痛部位",
        "疼痛等级",
        "主课数据",
        "一句复盘",
        "明日调整",
        "训练警报",
        "达成率",
    ],
    SHEET_REVIEWS: [
        "训练块名称",
        "计划km",
        "实际km",
        "完成率",
        "I有效km",
        "T1有效km",
        "T2有效km",
        "M有效km",
        "R短速度km",
        "平均RPE",
        "平均体重kg",
        "最高疼痛等级",
        "本周复盘",
        "下周调整",
    ],
    SHEET_PACE_RULES: ["代号", "类型", "目标配速", "生理目的", "备注", "排序"],
}

MAIN_TYPE_OPTIONS = ["REC", "E", "E+R", "LSD", "M", "T", "T1", "T2", "I", "I/R", "R", "Rest", "Mixed"]
BLOCK_TYPE_OPTIONS = ["week", "transition", "special"]
STATUS_OPTIONS = ["高质量完成", "一般完成", "降级完成", "没完成", "未完成", "休息", "取消/休息", "跳过"]
PAIN_LEVEL_OPTIONS = [str(value) for value in range(0, 11)]


def generate_excel_template_bytes() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _build_instructions_sheet(workbook)
    _build_data_sheet(
        workbook,
        SHEET_CYCLE,
        SHEET_HEADERS[SHEET_CYCLE],
        [
            [
                "2026夏训",
                "眉山东坡半马 1:11:30",
                "2026-06-01",
                "2026-09-06",
                "眉山东坡半马",
                "2026-11-08",
                "1:11:30",
                "夏季系统训练周期",
            ]
        ],
        comments={"周期名称": "一个 Excel 文件暂时只读取第一条有效训练周期。"},
    )
    _build_data_sheet(
        workbook,
        SHEET_BLOCKS,
        SHEET_HEADERS[SHEET_BLOCKS],
        [
            [
                "Week 1：重新启动周",
                "week",
                1,
                1,
                "6.1-6.7",
                88,
                94,
                92,
                "2026-06-01",
                "2026-06-07",
                "6月计划",
                "恢复接量，重新建立节奏",
            ]
        ],
        comments={"块类型": "允许值：week、transition、special。", "排序": "必填。"},
        validations={"块类型": BLOCK_TYPE_OPTIONS},
    )
    _build_data_sheet(
        workbook,
        SHEET_WORKOUTS,
        SHEET_HEADERS[SHEET_WORKOUTS],
        [
            [
                "2026-06-01",
                "周一",
                "Week 1：重新启动周",
                "6月计划",
                "轻松跑 10km + 4×100m",
                "控制心率，恢复节奏",
                10,
                "E",
                "4:45-5:30/km",
                1,
            ]
        ],
        comments={"训练块名称": "必须与“训练块”Sheet 中的训练块名称一致。", "主类型": "请选择下拉选项。"},
        validations={"主类型": MAIN_TYPE_OPTIONS},
    )
    _build_data_sheet(
        workbook,
        SHEET_LOGS,
        SHEET_HEADERS[SHEET_LOGS],
        [
            [
                "2026-06-01",
                "高质量完成",
                10.2,
                3000,
                "4:54",
                142,
                4,
                0,
                0,
                0,
                0,
                0.4,
                7.5,
                82,
                44,
                68.8,
                "轻松",
                "无",
                0,
                "4×100m 放松加速",
                "状态不错，控制得比较稳",
                "明天正常训练",
                "无",
                102,
            ]
        ],
        comments={
            "日期": "通过日期匹配训练计划。",
            "完成状态": "空值会按未开始处理。",
            "疼痛等级": "范围 0-10。",
        },
        validations={"完成状态": STATUS_OPTIONS, "疼痛等级": PAIN_LEVEL_OPTIONS},
    )
    _build_data_sheet(
        workbook,
        SHEET_REVIEWS,
        SHEET_HEADERS[SHEET_REVIEWS],
        [
            [
                "Week 1：重新启动周",
                92,
                88.5,
                96.2,
                5,
                8,
                0,
                0,
                1.2,
                5.5,
                68.8,
                1,
                "整体接量顺利，强度控制合理",
                "下周恢复正常二四日结构",
            ]
        ],
        comments={"训练块名称": "必须与“训练块”Sheet 中的训练块名称一致。", "最高疼痛等级": "范围 0-10。"},
        validations={"最高疼痛等级": PAIN_LEVEL_OPTIONS},
    )
    _build_data_sheet(
        workbook,
        SHEET_PACE_RULES,
        SHEET_HEADERS[SHEET_PACE_RULES],
        _default_pace_rule_rows(),
        comments={"代号": "同一用户下代号唯一，重复导入时会更新。"},
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_instructions_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(SHEET_INSTRUCTIONS)
    sheet["A1"] = "GaitLogic Planner 标准 Excel 模板填写说明"
    sheet["A1"].font = Font(bold=True, size=14)
    instructions = [
        "本模板用于导入 GaitLogic Planner 训练计划。",
        "请不要修改 Sheet 名称。",
        "请不要修改表头名称。",
        "日期格式建议使用 YYYY-MM-DD。",
        "距离单位为 km。",
        "时长字段建议使用秒，或使用 HH:MM:SS，后端兼容这两种格式。",
        "配速字段建议使用 秒/公里，或使用 mm:ss，后端兼容这两种格式。",
        "疼痛等级范围为 0-10。",
        "完成状态和训练类型请优先使用下拉选项。",
    ]
    for row_index, text in enumerate(instructions, start=3):
        sheet.cell(row=row_index, column=1, value=text)
    sheet.column_dimensions["A"].width = 88


def _build_data_sheet(
    workbook: Workbook,
    sheet_name: str,
    headers: list[str],
    rows: list[list[Any]],
    comments: dict[str, str] | None = None,
    validations: dict[str, list[str]] | None = None,
) -> None:
    sheet = workbook.create_sheet(sheet_name)
    sheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    comments = comments or {}
    validations = validations or {}

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        if header in comments:
            cell.comment = Comment(comments[header], "GaitLogic")
        sheet.column_dimensions[get_column_letter(column_index)].width = max(14, min(28, len(header) + 8))

    for row_index, row in enumerate(rows, start=2):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)

    for header, options in validations.items():
        column_index = headers.index(header) + 1
        column_letter = get_column_letter(column_index)
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(options)}"',
            allow_blank=True,
        )
        sheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}500")


def _default_pace_rule_rows() -> list[list[Any]]:
    return [
        ["R", "短速度", "100m放松快跑；200m 34-37秒；300m 50-53秒", "保持1500/5000速度火种", "快但放松，不是硬冲", 1],
        ["I", "间歇跑", "3:20-3:35/km", "提升最大摄氧与速度耐受", "按组间恢复控制质量", 2],
        ["T2", "高阈值", "3:38-3:45/km", "高阈值刺激", "偏强但可控", 3],
        ["T1", "稳阈值", "3:48-3:58/km", "稳定阈值能力", "更适合长节奏", 4],
        ["M", "稳态跑", "4:05-4:20/km", "提升马拉松强度耐受", "稳定输出", 5],
        ["E", "轻松跑", "4:45-5:30/km", "有氧积累", "控制心率", 6],
        ["REC", "恢复跑", "5:20-6:10/km", "恢复与低压力有氧", "轻松完成", 7],
        ["LSD", "长距离", "4:45-5:45/km", "长时间有氧与肌耐力", "注意补给", 8],
    ]
