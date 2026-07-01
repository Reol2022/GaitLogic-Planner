from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

TEMPLATE_FILENAME = "plan-import-template.xlsx"
HEADERS = [
    "date",
    "day_offset",
    "session_index",
    "workout_type",
    "title",
    "distance_km",
    "duration_min",
    "target_pace",
    "target_rpe",
    "content",
    "notes",
    "is_rest_day",
]


def generate_plan_import_template_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "plan_import"
    fill = PatternFill("solid", fgColor="DDEBF7")
    for index, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.font = Font(bold=True)
        cell.fill = fill
        sheet.column_dimensions[cell.column_letter].width = max(14, len(header) + 4)
    examples = [
        ["2026-07-10", None, 1, "E", "有氧跑", 12, None, "4:30-4:50/km", 4, "有氧跑 12km", "不追加训练量", False],
        [None, 1, 1, "I", "间歇", 14, None, "3:15/km", 8, "6x1000m，组间慢跑2分钟", "", False],
        [None, 2, 1, "REST", "休息", 0, None, "-", None, "休息", "", True],
    ]
    for row_index, row in enumerate(examples, start=2):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
