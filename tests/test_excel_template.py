from io import BytesIO

from openpyxl import load_workbook

from server.services.excel_template_service import (
    SHEET_HEADERS,
    STANDARD_SHEETS,
    generate_excel_template_bytes,
)


def test_generate_excel_template_bytes_can_open_workbook():
    content = generate_excel_template_bytes()

    assert content.startswith(b"PK")
    workbook = load_workbook(BytesIO(content), read_only=True)
    assert workbook.sheetnames == STANDARD_SHEETS


def test_excel_template_headers_are_standard():
    workbook = load_workbook(BytesIO(generate_excel_template_bytes()), read_only=True)

    for sheet_name, expected_headers in SHEET_HEADERS.items():
        sheet = workbook[sheet_name]
        headers = [sheet.cell(row=1, column=index).value for index in range(1, len(expected_headers) + 1)]
        assert headers == expected_headers
