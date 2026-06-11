from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from planner_core.enums import WorkoutMainTypeNormalized
from server.common.exceptions import BadRequestError
from server.services.ai_plan_export_service import export_ai_plan_draft


def make_draft():
    workouts = [
        SimpleNamespace(
            workout_date=date(2026, 6, 1),
            weekday="周一",
            block_name="Week 1",
            phase_name="基础期",
            planned_content="轻松跑 10km + 4x100m",
            focus_note="控制心率",
            planned_distance_km=Decimal("10.0"),
            main_type_raw="E",
            main_type_normalized=WorkoutMainTypeNormalized.easy,
            target_pace_text="4:40-5:20/km",
            sort_order=1,
        ),
        SimpleNamespace(
            workout_date=date(2026, 6, 2),
            weekday="周二",
            block_name="Week 1",
            phase_name="基础期",
            planned_content="阈值跑 4x2km",
            focus_note="可控阈值",
            planned_distance_km=Decimal("14.0"),
            main_type_raw="T1",
            main_type_normalized=WorkoutMainTypeNormalized.tempo,
            target_pace_text="3:35-3:45/km",
            sort_order=2,
        ),
    ]
    return SimpleNamespace(
        id=3,
        title="8 周半马计划",
        goal="半马 1:11:30",
        start_date=date(2026, 6, 1),
        end_date=date(2026, 7, 26),
        target_race_name="测试半马",
        target_race_date=date(2026, 8, 1),
        target_result="1:11:30",
        summary="稳步推进，控制强度。",
        risk_notes=["注意恢复"],
        status="draft",
        created_at=datetime(2026, 6, 1),
        updated_at=datetime(2026, 6, 1),
        workouts=workouts,
    )


def test_export_ai_plan_draft_xlsx_can_open() -> None:
    export_file = export_ai_plan_draft(make_draft(), "xlsx")
    workbook = load_workbook(BytesIO(export_file.content), read_only=True)
    assert "训练计划" in workbook.sheetnames
    sheet = workbook["训练计划"]
    assert sheet["A1"].value == "日期"
    assert sheet["E2"].value == "轻松跑 10km + 4x100m"


def test_export_ai_plan_draft_csv_and_markdown() -> None:
    csv_file = export_ai_plan_draft(make_draft(), "csv")
    assert csv_file.filename.endswith(".csv")
    assert "轻松跑" in csv_file.content.decode("utf-8-sig")

    markdown_file = export_ai_plan_draft(make_draft(), "markdown")
    markdown = markdown_file.content.decode("utf-8")
    assert markdown.startswith("# 8 周半马计划")
    assert "| 2026-06-01 | 周一 |" in markdown


def test_export_ai_plan_draft_device_csv_and_ics() -> None:
    device_file = export_ai_plan_draft(make_draft(), "garmin_csv")
    device_csv = device_file.content.decode("utf-8-sig")
    assert "Garmin manual reference" in device_csv
    assert "workout_description" in device_csv

    ics_file = export_ai_plan_draft(make_draft(), "ics")
    ics = ics_file.content.decode("utf-8")
    assert "BEGIN:VCALENDAR" in ics
    assert "SUMMARY:E 10km" in ics


def test_export_ai_plan_draft_rejects_unknown_format() -> None:
    with pytest.raises(BadRequestError):
        export_ai_plan_draft(make_draft(), "fit")
